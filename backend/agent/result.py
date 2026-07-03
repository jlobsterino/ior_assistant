"""
result - сборка «result-пакета» выгрузки для премиального UI.

Аудитору нужна не просто кнопка-файл, а доказательство, которое можно защитить и
повторить. Поэтому из того, что бэкенд уже знает (spec_resolved, lineage-воронка,
финальный df), собираем:
* conditions  - применённые условия человеческим языком + точно (колонки) -> чипы;
* methodology - как именно посчитано (текст для дела);
* funnel      - воронка строк по этапам (3 418 -> 457 -> 24);
* preview     - топ-строки результата (таблица в чате);
* summary     - ключевые числа (всего / максимум / сумма метрики);
* confidence  - где агент додумал / нестандартная методология.

Чистый офлайн-безопасный модуль: pandas импортируется ЛЕНИВО внутри функций,
работающих с df. Опирается на status.py (человеческие имена/форматы).
"""
from __future__ import annotations

from typing import Optional

from backend.agent.status import fmt_int, fmt_money, human_column, human_table

_OP = {"gt": ">", "gte": "≥", "lt": "<", "lte": "≤", "ne": "≠", "eq": "=",
       ">": ">", ">=": "≥", "<": "<", "<=": "≤", "=": "="}

# суммовые алиасы, которые форматируем как деньги
_MONEY_HINT = ("loss", "sum", "amt", "потер", "возмещ", "recovery", "net")


def _is_money_name(name: str) -> bool:
    n = str(name or "").lower()
    return any(h in n for h in _MONEY_HINT)


def build_conditions(spec: dict) -> list:
    """Применённые условия -> список чипов {kind,label,detail,column,editable}."""
    if not isinstance(spec, dict):
        return []
    out: list = []
    labels = ((spec.get("period") or {}).get("labels") or {})
    for f in (spec.get("filters") or []):
        if not isinstance(f, dict):
            continue
        kind = (f.get("kind") or "").lower()
        col = f.get("column")
        if kind == "period":
            intent = f.get("intent")
            txt = (intent.get("text") if isinstance(intent, dict) else intent) or ""
            out.append({"kind": "period", "label": "Период",
                        "detail": labels.get(col) or str(txt), "column": col,
                        "editable": True})
        elif kind == "categorical":
            out.append({"kind": "filter", "label": human_column(col) or "Фильтр",
                        "detail": str(f.get("value") or ""), "column": col,
                        "editable": True})
        elif kind == "range":
            op = _OP.get((f.get("op") or "gt").lower(), ">")
            val = f.get("value")
            d = f"{op} {fmt_money(val) if _is_money_name(col) else fmt_int(val)}"
            out.append({"kind": "range", "label": col, "detail": d,
                        "column": col, "editable": True})
        elif kind == "like":
            out.append({"kind": "text", "label": "Текст",
                        "detail": str(f.get("value") or "").strip("%"),
                        "column": col, "editable": True})
    joins = (spec.get("source") or {}).get("joins") or []
    if joins:
        out.append({"kind": "join", "label": "Связано",
                    "detail": ", ".join(human_table(j.get("table"))
                                        for j in joins if isinstance(j, dict)),
                    "editable": False})
    agg = spec.get("aggregate") or {}
    if agg.get("group_by"):
        out.append({"kind": "group", "label": "Группировка",
                    "detail": ", ".join(human_column(c) for c in agg["group_by"]),
                    "editable": True})
    dm = [m.get("as") for m in (spec.get("derived_metrics") or []) if m.get("as")]
    if dm:
        out.append({"kind": "metric", "label": "Метрика",
                    "detail": ", ".join(dm), "editable": False})
    return out


def build_methodology(spec: dict) -> str:
    """Человеческий абзац «как посчитано» – для рабочего дела аудитора."""
    if not isinstance(spec, dict):
        return ""
    parts: list[str] = []
    conds = build_conditions(spec)
    period = next((c for c in conds if c["kind"] == "period"), None)
    if period:
        parts.append(f"Период: {period['detail']}.")
    cats = [c for c in conds if c["kind"] in ("filter", "range", "text")]
    if cats:
        parts.append("Условия: " + "; ".join(f"{c['label']} {c['detail']}"
                                             if c["kind"] == "range"
                                             else f"{c['label']} = {c['detail']}"
                                             for c in cats) + ".")
    joins = (spec.get("source") or {}).get("joins") or []
    if joins:
        money_via = ", ".join(human_table(j.get("table")) for j in joins if isinstance(j, dict))
        parts.append(f"Денежные показатели взяты через присоединение таблиц «{money_via}» "
                     f"(агрегация по incdnt_id), т.к. суммовые поля основной таблицы "
                     f"заполнены лишь ~2.26%.")
    agg = spec.get("aggregate") or {}
    if agg.get("group_by"):
        parts.append("Группировка по " + ", ".join(human_column(c) for c in agg["group_by"]) + ".")
    dm = spec.get("derived_metrics") or []
    for m in dm:
        expr = m.get("expr") or {}
        if expr.get("op") == "sub":
            parts.append(f"{m.get('as')} = {expr.get('left')} - {expr.get('right')}.")
    return " ".join(parts)


def build_funnel(funnel: list) -> list:
    """Воронка строк: [{stage, rows}] -> очищенный список (без None-rows)."""
    out = []
    for f in (funnel or []):
        if isinstance(f, dict) and f.get("rows") is not None:
            out.append({"stage": f.get("stage"), "rows": int(f["rows"])})
    return out


def build_preview(df, n: int = 8) -> dict:
    """{headers, rows} – топ-n строк финального df строками (для таблицы в чате)."""
    try:
        cols = list(df.columns)[:12]
        headers = [human_column(c) if human_column(c) != c else str(c) for c in cols]
        rows = []
        for _, r in df.head(n).iterrows():
            rows.append([_cell(r[c]) for c in cols])
        return {"headers": headers, "raw_headers": [str(c) for c in cols],
                "rows": rows, "total": int(len(df)),
                "truncated": int(len(df)) > n}
    except Exception:  # noqa: BLE001
        return {"headers": [], "rows": [], "total": 0, "truncated": False}


def build_summary(df, spec: dict) -> dict:
    """Ключевые числа: всего + (для агрегата) максимум/сумма главной метрики."""
    import math
    total = int(len(df)) if df is not None else 0
    agg = (spec or {}).get("aggregate") or {}
    is_agg = bool(agg.get("group_by"))
    highlights: list = []
    if total == 0:
        return {"total": 0, "is_aggregate": is_agg, "highlights": []}

    # главная метрика: последний derived_metric, иначе первая sum-метрика агрегата
    metric = None
    dm = (spec or {}).get("derived_metrics") or []
    if dm:
        metric = dm[-1].get("as")
    if not metric:
        for m in (agg.get("metrics") or []):
            if m.get("fn") == "sum" and m.get("as"):
                metric = m["as"]; break

    gb = (agg.get("group_by") or [None])[0]
    try:
        if is_agg and metric and metric in df.columns and gb and gb in df.columns:
            ser = df[metric]
            money = _is_money_name(metric)
            total_v = float(ser.sum())
            highlights.append({"label": f"Всего {human_column(gb)}", "value": fmt_int(total)})
            idx = ser.idxmax()
            top_grp = str(df.loc[idx, gb])
            top_v = float(ser.loc[idx])
            highlights.append({
                "label": "Максимум", "value": (fmt_money(top_v) if money else fmt_int(top_v)),
                "sub": top_grp[:48]})
            highlights.append({
                "label": "Суммарно", "value": fmt_money(total_v) if money else fmt_int(total_v)})
        else:
            highlights.append({"label": "Строк", "value": fmt_int(total)})
    except Exception:  # noqa: BLE001
        highlights = [{"label": "Строк", "value": fmt_int(total)}]
    return {"total": total, "is_aggregate": is_agg, "metric": metric,
            "highlights": highlights}


def build_bars(df, spec: dict, n: int = 8) -> Optional[dict]:
    """Мини-бар топ-N для агрегата: {label_col, value_col, items:[{label,value,raw}]}."""
    agg = (spec or {}).get("aggregate") or {}
    if not agg.get("group_by") or df is None or len(df) == 0:
        return None
    gb = agg["group_by"][0]
    dm = (spec or {}).get("derived_metrics") or []
    metric = dm[-1].get("as") if dm else None
    if not metric:
        for m in (agg.get("metrics") or []):
            if m.get("fn") == "sum":
                metric = m.get("as"); break
    if not metric or metric not in df.columns or gb not in df.columns:
        return None

    try:
        top = df.sort_values(metric, ascending=False).head(n)
        money = _is_money_name(metric)
        mx = float(top[metric].max()) or 1.0
        items = []
        for _, r in top.iterrows():
            v = float(r[metric])
            items.append({"label": str(r[gb])[:40],
                          "value": fmt_money(v) if money else fmt_int(v),
                          "pct": max(0.0, min(1.0, v / mx)) if mx else 0.0})
        return {"label_col": human_column(gb), "value_col": metric, "items": items}
    except Exception:  # noqa: BLE001
        return None


def _cell(v) -> str:
    import math
    if v is None:
        return ""
    try:
        if isinstance(v, float):
            if math.isnan(v):
                return ""
            if v.is_integer():
                v = int(v)
        if isinstance(v, (int, float)):
            return fmt_int(v) if abs(v) >= 1000 else str(v)
        if hasattr(v, "isoformat"):
            return v.isoformat(sep=" ")[:16]
        s = str(v)
        return s if len(s) <= 80 else s[:77] + "..."
    except Exception:  # noqa: BLE001
        return str(v)[:80]


def write_methodology_sheet(path: str, spec: dict, funnel: list = None) -> bool:
    """Добавляет лист «Методология» в готовый xlsx (П5) – для рабочего дела аудитора:
    применённые условия + воронка + текст «как посчитано». Возвращает успех."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font
        wb = load_workbook(path)
        if "Методология" in wb.sheetnames:
            del wb["Методология"]
        ws = wb.create_sheet("Методология")
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 80
        ws["A1"] = "Методология выгрузки"
        ws["A1"].font = Font(bold=True, size=13)
        r = 3
        ws.cell(row=r, column=1, value="Условие").font = Font(bold=True)
        ws.cell(row=r, column=2, value="Значение").font = Font(bold=True)
        r += 1
        for c in build_conditions(spec):
            ws.cell(row=r, column=1, value=c.get("label"))
            ws.cell(row=r, column=2, value=str(c.get("detail") or ""))
            r += 1
        for f in build_funnel(funnel):
            ws.cell(row=r, column=1, value=f"Этап: {f.get('stage')}")
            ws.cell(row=r, column=2, value=fmt_int(f.get("rows")))
            r += 1
        r += 1
        ws.cell(row=r, column=1, value="Как посчитано").font = Font(bold=True)
        cell = ws.cell(row=r, column=2, value=build_methodology(spec))
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=2)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb.save(path)
        return True
    except Exception:  # noqa: BLE001
        return False


def build_result_package(spec: dict, df, funnel: list = None,
                         warnings: list = None, file_id=None) -> dict:
    """Полный пакет для UI-события `result`."""
    return {
        "conditions": build_conditions(spec),
        "methodology": build_methodology(spec),
        "funnel": build_funnel(funnel),
        "preview": build_preview(df),
        "summary": build_summary(df, spec),
        "bars": build_bars(df, spec),
        "warnings": list(warnings or []),
        "spec": spec,
        "file_id": file_id,
    }