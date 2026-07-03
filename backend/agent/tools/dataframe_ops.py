"""
Atomic tools для работы с dataframes и витриной БЗ.

  * query         - Spark SQL по таблице БЗ -> df_id
  * filter_df     - pandas .query() на существующем df -> df_id
  * top_n         - sort + head -> df_id
  * group_by      - groupby + агрегаты -> df_id
  * join_dfs      - merge двух df'ов -> df_id
  * export_excel  - xlsx из df -> file_id
  * export_csv    - csv из df -> file_id
  * get_ior_details - досье одного ИОР (composite: query + 3 joins)

Все tools работают с pandas (после Spark.toPandas()) - это ОК для размеров
которые мы возвращаем в чат (limit 100k строк по умолчанию).
"""

from __future__ import annotations

import asyncio
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from backend.agent.schema import get_schema
from backend.agent.tools.base import Tool, ToolResult
from backend.agent.tools.registry import REGISTRY
from backend.config import get_settings
from backend.data import get_data_store

logger = logging.getLogger(__name__)

_MAX_ROWS_DEFAULT = 100_000
_MAX_ROWS_HARD = 2_000_000


def _money_main_warning(cols: list) -> Optional[str]:
    """Мягкое предупреждение (§2.4-bis): агрегат/сортировка по money-колонке main
    (filled_pct<порог, метаданные - НЕ подстрока имени). Тул не блокирует (не всегда
    знает источник df), лишь подсказывает «деньги через join к fin_impact/recovery».
    Жесткий blocker - в validate_spec, где источник колонки известен точно.
    """
    from backend.agent.query_spec import is_money_main_col  # lazy: pydantic-safe
    schema = get_schema()
    hit = [c for c in cols if isinstance(c, str) and is_money_main_col(c, schema)]
    if not hit:
        return None
    return (f"▲ {hit} — суммовые колонки main заполнены ~2.26%, агрегат будет почти "
            f"\"пустым. Бери деньги через join к fin_impact/recovery.\")")


# —— query: Spark SQL по таблице БЗ ------------------------------------


def _validate_enum_filters(table: str, where: Optional[dict]) -> Optional[str]:
    """Если в where есть `col = value` для колонки с enum_values, и value
    не в enum — возвращаем подсказку (ошибка для reflector'а).
    Не валидируем !=, LIKE, ранжи — только прямые equality.

    LAZY ENRICH: если колонка-кандидат (по эвристике) ещё не имеет
    enum_values, пытаемся загрузить их SELECT DISTINCT прямо сейчас.
    Это покрывает редко-используемые категориальные колонки которые
    при startup-enrich не попали (если их там >max_columns).
    """
    if not where:
        return None
    schema = get_schema()
    t = schema.get(table)
    if t is None:
        return None

    # Lazy enrich: для каждой колонки из where, на которую модель сделала
    # равенство-фильтр, проверяем - это enum-кандидат? Если да и ещё не
    # загружено - пробуем подгрузить из БД (1 запрос).
    from backend.agent.schema.loader import _is_enum_candidate, enrich_one_column
    from backend.data import get_data_store
    for key in where:
        if "__" in key:
            continue
        col = next((c for c in t.columns if c.name == key), None)
        if col is None or col.enum_values is not None:
            continue
        if not _is_enum_candidate(col):
            continue
        try:
            enrich_one_column(get_data_store(), table, key)
        except Exception:  # noqa: BLE001
            pass  # игнор - просто продолжим без enum-валидации

    enum_map = {c.name: c.enum_values for c in t.columns if c.enum_values}
    if not enum_map:
        return None
    for key, val in where.items():
        # Только прямые =. Скипаем __like / __ne / __gte / dict-ranges / list (IN).
        if "__" in key or isinstance(val, (dict, list)) or val is None:
            continue
        if key in enum_map and isinstance(val, str) and val not in enum_map[key]:
            valid = ", ".join(f"'{v}'" for v in enum_map[key])
            hint = ""
            v_lower = val.lower()
            if any(s in v_lower for s in ("не закрыт", "незакрыт", "открыт", "active", "open")):
                hint = (" Семантика «незакрытый» = НЕ 'Закрыт' — используй "
                        f"\"\\\"{key}__ne\\\": \\\"Закрыт\\\" (или filter_df со строкой "
                        f"\\\"\\\"{key}\\\" != 'Закрыт'\\\").")
            return (f"ENUM-нарушение: {key}={val!r} не валидное значение. "
                    f"Допустимые: [{valid}].{hint}")
    return None


async def query(ctx, table: str, where: Optional[dict] = None,
                columns: Optional[list] = None,
                limit: int = _MAX_ROWS_DEFAULT,
                order_by: Optional[str] = None,
                order_desc: bool = True,
                # LLM-friendly aliases (qwen/giga любят свои имена):
                fields: Optional[list] = None,
                select: Optional[list] = None,
                sort_by: Optional[str] = None) -> ToolResult:
    """SELECT из одной таблицы БЗ с фильтрами/проекцией/limit.

    Идёт через `IORDataStore` (Spark в prod / DuckDB в local) - один и
    тот же код работает в обеих средах.
    """
    # Алиасы для проекции
    columns = columns or fields or select
    # LLM иногда передаёт [""] вместо null - нормализуем в "взять все"
    if columns and len(columns) == 1 and columns[0] in ("", "all", "*"):
        columns = None
    # Чисто-пустой list тоже -> all
    if columns is not None and len(columns) == 0:
        columns = None

    # Алиас для order_by
    if order_by is None and sort_by is not None:
        order_by = sort_by

    # Нормализация order_by: LLM может прислать list/dict/int - приводим к str|None
    if isinstance(order_by, dict):
        # {"col": "desc"} или {"col": -1}
        if order_by:
            k, v = next(iter(order_by.items()))
            order_by = k if isinstance(k, str) else None
            if isinstance(v, str):
                order_desc = (v.lower() in ("desc", "-1", "descending"))
            else:
                order_desc = (v == -1 or v is False)
        else:
            order_by = None
    elif isinstance(order_by, list):
        # list of strings - берём первый. list из int/чего-то ещё -> drop.
        first = order_by[0] if order_by else None
        order_by = first if isinstance(first, str) else None
    elif order_by is not None and not isinstance(order_by, str):
        # int/float/bool - это не имя колонки. Дропаем (top_n шаг сделает сортировку).
        logger.warning("[query] order_by=%r не строка - игнорирую", order_by)
        order_by = None

    # Django-style "-col" -> DESC
    if isinstance(order_by, str) and order_by.startswith("-"):
        order_by = order_by[1:]
        order_desc = True
    elif isinstance(order_by, str) and order_by.startswith("+"):
        order_by = order_by[1:]
        order_desc = False

    # LLM иногда передаёт limit=null или вообще не передаёт - фоллбэк на default
    if limit is None:
        limit = _MAX_ROWS_DEFAULT
    try:
        limit = min(int(limit), _MAX_ROWS_HARD)
    except (TypeError, ValueError):
        limit = _MAX_ROWS_DEFAULT

    # Pre-validate ENUM values: ловим hallucinations типа 'Не закрыт' до SQL.
    enum_err = _validate_enum_filters(table, where)
    if enum_err:
        return ToolResult(ok=False, error=enum_err)

    store = get_data_store()
    try:
        # Spark.sql().toPandas() - БЛОКИРУЮЩИЙ вызов (5-30 секунд в prod).
        # Без to_thread весь event loop стоит -> WS event'ы не доходят
        # до фронта, executor зависает посреди шага. В local с DuckDB
        # это микросекунды и проблемы не было.
        df = await asyncio.to_thread(
            store.query,
            table=table, where=where, columns=columns,
            order_by=order_by, order_desc=order_desc, limit=limit,
        )
    except ValueError as e:
        return ToolResult(ok=False, error=str(e))
    except FileNotFoundError as e:
        return ToolResult(ok=False, error=str(e))
    except Exception as e:  # noqa: BLE001
        return ToolResult(
            ok=False,
            error=f"query упал: {type(e).__name__}: {e}",
        )

    # —— ГАРД «без молчаливого дампа» ———————————————————————————————————
    # 0 строк при наличии фильтра - НЕ ok. Раньше это молча уходило в Excel
    # (корневой баг тестировщиков: фильтр на неверную колонку -> 0 -> дамп).
    # Диагностируем по реальному каталогу и отдаём рефлектору actionable-ошибку
    # (какую колонку поставить / какое реальное значение взять).
    if where and len(df) == 0:
        from backend.agent.resolve.grounding import diagnose_empty
        try:
            diag = diagnose_empty(table, where)
            return ToolResult(ok=False, error=diag.message)
        except Exception as e:  # noqa: BLE001
            logger.warning("[query] diagnose_empty упал: %s", e)
        return ToolResult(
            ok=False,
            error=(f"EMPTY_RESULT: query({table}) вернул 0 строк по фильтуру "
                   f"'{where}'. Проверь колонку/значение через search_values "
                   f"или верни ask_user."))

    desc = f"query({table}), where={where}) -> {len(df)} rows"
    meta = ctx.register_dataframe(df, description=desc, created_by="query")
    return ToolResult(
        ok=True,
        output={"df_id": meta.df_id, "rows": meta.rows,
                "columns": meta.columns, "sample": meta.sample},
        summary=f"'{meta.df_id}': {meta.rows} rows * {len(meta.columns)} cols "
                f"из '{table}'",
    )


async def _resolve_df(ctx, df_id: Optional[str]) -> tuple[pd.DataFrame, str]:
    """Helper to resolve df_id from ctx.dataframes, table names, or defaults."""
    import pandas as pd
    from backend.agent.schema import get_schema
    from backend.data import get_data_store
    
    # 1. If df_id is empty/None
    if not df_id:
        if len(ctx.dataframes) == 1:
            df_id = list(ctx.dataframes.keys())[0]
        else:
            df_id = "d6_base_of_knowledge_ior"
            
    # 2. If df_id is a table name in the database
    schema = get_schema()
    is_table = df_id in schema.table_names() or (isinstance(df_id, str) and df_id.startswith("d6_"))
    if is_table:
        if df_id in ctx.dataframes:
            return ctx.dataframes[df_id], df_id
        # Load from DB
        store = get_data_store()
        df = await asyncio.to_thread(store.query, table=df_id)
        meta = ctx.register_dataframe(df, description=f"auto-loaded {df_id}", created_by="auto-loader")
        return df, meta.df_id
        
    # 3. If df_id is not in dataframes, but we have only one dataframe registered, use it!
    if df_id not in ctx.dataframes and len(ctx.dataframes) == 1:
        df_id = list(ctx.dataframes.keys())[0]
        
    # 4. Standard lookup (will raise KeyError if still not found)
    return ctx.get_df(df_id), df_id


async def filter_df(ctx, df_id: str, where: str) -> ToolResult:
    """Pandas-фильтр через DataFrame.query(). Возвращает новый df_id.

    `where` - строка вида "incdnt_sum > 1000000 and incdnt_status_name != 'Закрыт'"
    Поддерживает синтаксис pandas .query().
    """
    df, df_id = await _resolve_df(ctx, df_id)
    try:
        # pandas .query() блокирующий — на больших df даёт заметный freeze
        filtered = await asyncio.to_thread(df.query, where, engine="python")
    except Exception as e:
        return ToolResult(
            ok=False,
            error=f"filter упал: {type(e).__name__}: {e}\nwhere: {where!r}",
        )
    desc = f"filter({df_id}, {where[:80]}...) -> {len(filtered)} rows"
    meta = ctx.register_dataframe(filtered, description=desc,
                                  created_by="filter_df")
    return ToolResult(ok=True, output={"df_id": meta.df_id, "rows": meta.rows},
                      summary=f"'{meta.df_id}': {meta.rows} rows (was {len(df)})")


async def top_n(ctx, df_id: str, by: str, n: int = 10,
                ascending: bool = False) -> ToolResult:
    """Топ-N строк по полю `by`."""
    df, df_id = await _resolve_df(ctx, df_id)
    if by not in df.columns:
        return ToolResult(
            ok=False,
            error=f"top_n: колонки {by!r} нет в {df_id}. "
                  f"Есть: {list(df.columns)[:20]}",
        )
    sorted_df = await asyncio.to_thread(
        lambda: df.sort_values(by=by, ascending=ascending,
                               na_position="last").head(n),
    )
    desc = f"top_{n}({df_id}), by={by}, {'asc' if ascending else 'desc'})"
    meta = ctx.register_dataframe(sorted_df, description=desc,
                                  created_by="top_n")
    summary = f"'{meta.df_id}': топ-{n} из {len(df)} по {by}"
    mw = _money_main_warning([by])
    if mw:
        summary += " " + mw
    return ToolResult(ok=True, output={"df_id": meta.df_id, "rows": meta.rows},
                      summary=summary)


async def group_by(ctx, df_id: str, by: list, agg: dict) -> ToolResult:
    """Group + агрегаты. agg = {column: 'sum'|'mean'|'count'|'max'|'min'}.

    Пример: group_by(df_id, by=['org_struct_lvl1_2_name'],
                     agg={'incdnt_sum': 'sum', 'incdnt_id': 'count'})
    """
    df, df_id = await _resolve_df(ctx, df_id)
    bad = [c for c in by if c not in df.columns] + \
          [c for c in agg if c not in df.columns]
    if bad:
        return ToolResult(
            ok=False,
            error=f"group_by: колокнок {bad} нет в {df_id}",
        )
    try:
        grouped = await asyncio.to_thread(
            lambda: df.groupby(by, dropna=False).agg(agg).reset_index(),
        )
    except Exception as e:
        return ToolResult(
            ok=False,
            error=f"group_by упал: {type(e).__name__}: {e}",
        )
    desc = f"group_by({df_id}, by={by}, agg={agg})"
    meta = ctx.register_dataframe(grouped, description=desc,
                                  created_by="group_by")
    summary = f"'{meta.df_id}': {meta.rows} групп"
    mw = _money_main_warning([c for c in agg])
    if mw:
        summary += " " + mw
    return ToolResult(ok=True,
                      output={"df_id": meta.df_id, "rows": meta.rows,
                              "columns": list(grouped.columns)},
                      summary=summary)


async def derive_column(ctx, df_id: str, source: str, new_column: str,
                        op: str = "month") -> ToolResult:
    """Вычисляемое поле из существующего. Нужно для помесячных/поквартальных
    отчётов — извлечь месяц/год/квартал из даты ПЕРЕД group_by/window_rank.

    op:
      'year'         -> '2025'         (год)
      'month'        -> '2025-03'      (год-месяц, для помесячной группировки)
      'quarter'      -> '2025-Q1'      (год-квартал)
      'day'          -> '2025-03-15'   (дата без времени)
      'month_num'    -> '03'           (номер месяца без года)
    """
    df, df_id = await _resolve_df(ctx, df_id)
    if source not in df.columns:
        return ToolResult(
            ok=False,
            error=f"derive_column: колонки source={source!r} нет в {df_id}. "
                  f"Есть: {list(df.columns)[:20]}",
        )

    def _compute():
        import pandas as pd
        d = df.copy()
        s = pd.to_datetime(d[source], errors="coerce")
        if op == "year":
            d[new_column] = s.dt.year.astype("Int64").astype(str)
        elif op == "month":
            d[new_column] = s.dt.to_period("M").astype(str)
        elif op == "quarter":
            d[new_column] = (s.dt.year.astype("Int64").astype(str)
                             + "-Q" + s.dt.quarter.astype("Int64").astype(str))
        elif op == "day":
            d[new_column] = s.dt.date.astype(str)
        elif op == "month_num":
            d[new_column] = s.dt.month.astype("Int64").astype(str).str.zfill(2)
        else:
            raise ValueError(f"op={op!r} не поддерживается "
                             f"(year|month|quarter|day|month_num)")
        return d

    try:
        result = await asyncio.to_thread(_compute)
    except Exception as e:  # noqa: BLE001
        return ToolResult(ok=False,
                          error=f"derive_column упал: {type(e).__name__}: {e}")
    meta = ctx.register_dataframe(
        result, description=f"derive({df_id}, {new_column}={op}({source}))",
        created_by="derive_column",
    )
    return ToolResult(
        ok=True,
        output={"df_id": meta.df_id, "rows": meta.rows},
        summary=f"'{meta.df_id}': +колонка {new_column} ({op} из {source})",
    )


async def window_rank(ctx, df_id: str,
                      partition_by: Optional[str | list] = None,
                      order_by: str = None,
                      order_desc: bool = True,
                      top_n: Optional[int] = None,
                      rank_col: str = "rank",
                      method: str = "row_number",
                      # LLM aliases
                      partition: Optional[str | list] = None,
                      by: Optional[str] = None) -> ToolResult:
    """Оконная функция: ранжирование внутри групп (как SQL ROW_NUMBER/RANK
    OVER PARTITION BY ... ORDER BY ...).

    Нужно для отчётов вида «ТОП-20 ЦПР по количеству в КАЖДОМ месяце»,
    «аутсайдеры», «N-й по величине внутри группы».

    Args:
        partition_by: колонка(и) разбиения окна (например 'month').
                      Если None — одно глобальное окно.
        order_by: колонка сортировки внутри окна (например 'incdnt_count').
        order_desc: True = по убыванию (топ), False = по возрастанию (аутсайдеры).
        top_n: если задан — оставить только rank <= top_n внутри каждого окна.
        rank_col: имя колонки с рангом в результате (default 'rank').
        method: 'row_number' (уникальный ранг) | 'rank' (с пропусками при
                равенстве) | 'dense_rank' (без пропусков).
    """
    df, df_id = await _resolve_df(ctx, df_id)
    # Алиасы
    if partition_by is None:
        partition_by = partition
    if order_by is None:
        order_by = by
    if not order_by:
        return ToolResult(ok=False,
                          error="window_rank: нужен order_by (колонка сортировки)")
    if order_by not in df.columns:
        return ToolResult(
            ok=False,
            error=f"window_rank: колонки order_by={order_by!r} нет в {df_id}. "
                  f"Есть: {list(df.columns)[:20]}"
        )
    # partition_by -> список
    parts: list = []
    if partition_by:
        parts = [partition_by] if isinstance(partition_by, str) else list(partition_by)
        bad = [c for c in parts if c not in df.columns]
        if bad:
            return ToolResult(ok=False,
                              error=f"window_rank: колонок partition_by={bad} нет в {df_id}")

    pandas_method = {
        "row_number": "first",    # first -> уникальные последовательные ранги
        "rank": "min",
        "dense_rank": "dense",
    }.get(method, "first")

    def _compute():
        d = df.copy()
        ascending = not order_desc
        if parts:
            grp = d.groupby(parts, dropna=False)[order_by]
            d[rank_col] = grp.rank(method=pandas_method, ascending=ascending).astype(int)
        else:
            d[rank_col] = d[order_by].rank(method=pandas_method,
                                           ascending=ascending).astype(int)
        if top_n is not None:
            d = d[d[rank_col] <= int(top_n)]
        # Сортируем для читаемости: по partition, затем по rank
        sort_cols = parts + [rank_col] if parts else [rank_col]
        d = d.sort_values(sort_cols).reset_index(drop=True)
        return d

    try:
        result = await asyncio.to_thread(_compute)
    except Exception as e:  # noqa: BLE001
        return ToolResult(ok=False,
                          error=f"window_rank упал: {type(e).__name__}: {e}")

    pdesc = f"partition={parts}" if parts else "global"
    desc = (f"window_rank({df_id}, order={order_by} "
            f"{'desc' if order_desc else 'asc'}, top_n={top_n})")
    meta = ctx.register_dataframe(result, description=desc,
                                  created_by="window_rank")

    return ToolResult(
        ok=True,
        output={"df_id": meta.df_id, "rows": meta.rows,
                "columns": list(result.columns)},
        summary=(f"{meta.df_id}: {meta.rows} строк "
                 f"(ранжировано по {order_by}"
                 + (f", топ-{top_n} в окне" if top_n else "") + ")"),
    )


async def join_dfs(ctx, left_df: Optional[str] = None,
                   right_df: Optional[str] = None,
                   on: Optional[str | list] = None,
                   how: str = "inner",
                   allow_fanout: bool = False,
                   # LLM-aliases
                   left: Optional[str] = None,
                   right: Optional[str] = None,
                   key: Optional[str | list] = None,
                   left_on: Optional[str] = None,
                   right_on: Optional[str] = None) -> ToolResult:
    """Merge двух DF'ов по ключу. how: inner|left|right|outer.

    Аргументы: left_df + right_df + on. Алиасы: left/right, key=on,
    left_on/right_on (если ключи в DF разные).
    """
    # LLM-aliases
    left_df = left_df or left
    right_df = right_df or right
    on = on if on is not None else key
    if not left_df or not right_df:
        avail = list(ctx.dataframes.keys()) if hasattr(ctx, "dataframes") else []
        return ToolResult(
            ok=False,
            error=(f"join_dfs: нужны left_df + right_df (df_id). "
                   f"Доступно в сессии: {avail or '(пусто — сначала query)'}."),
        )

    try:
        a, left_df = await _resolve_df(ctx, left_df)
        b, right_df = await _resolve_df(ctx, right_df)
    except KeyError as e:
        return ToolResult(ok=False, error=str(e))
    if how not in ("inner", "left", "right", "outer"):
        return ToolResult(ok=False, error=f"how={how!r} invalid")

    # — ГАРД fan-out 1:N (§2.4-bis, variant 4) —
    # для how∈{left,inner} без allow_fanout: ключ on не уникален в правом df ->
    # строки взорвутся, деньги посчитаются дважды.
    if how in ("left", "inner") and on is not None and not (left_on and right_on) \
            and not allow_fanout:
        from backend.agent.query_spec import detect_fanout  # lazy: pydantic-safe
        if detect_fanout(b, on):
            return ToolResult(ok=False, error=(
                f"right_df не уникален по {on} (fan-out, двойной счёт) — агрегируй "
                f"right_df по ключу (group_by sum/count) или передай allow_fanout=true, "
                f"если это осознанно."))

    try:
        if left_on and right_on:
            merged = await asyncio.to_thread(
                a.merge, b, left_on=left_on, right_on=right_on, how=how)
        else:
            merged = await asyncio.to_thread(a.merge, b, on=on, how=how)
    except KeyError as e:
        # Helpful error: показать что есть в обеих сторонах
        lcols = list(a.columns)[:30]
        rcols = list(b.columns)[:30]
        return ToolResult(
            ok=False,
            error=(f"join упал: ключ {e} не найден. \n"
                   f"\"left_df\"={left_df} columns: {lcols}. \n"
                   f"\"right_df\"={right_df} columns: {rcols}. \n"
                   f"Подбери общий ключ или используй left_on+right_on."),
        )
    except Exception as e:  # noqa: BLE001
        return ToolResult(ok=False, error=f"join упал: {type(e).__name__}: {e}")

    desc = f"join({left_df}, {right_df}, on={on}, how={how}) -> {len(merged)} rows"
    meta = ctx.register_dataframe(merged, description=desc, created_by="join_dfs")
    summary = f"{meta.df_id}: {meta.rows} rows (left={len(a)}, right={len(b)})"
    
    # post-merge sanity: строки взорвались при left/inner — предупреждаем (allow_fanout)
    if how in ("left", "inner") and len(merged) > len(a):
        summary += (f" ▲ fan-out: строк стало больше left ({len(a)}->{len(merged)}) "
                    f"— возможен двойной счёт сумм.")
                    
    return ToolResult(
        ok=True,
        output={"df_id": meta.df_id, "rows": meta.rows},
        summary=summary,
    )


async def export_excel(ctx, df_id: str,
                       name: Optional[str] = None,
                       sheet_name: str = "Отчет_ОпРиски") -> ToolResult:
    """Сохранить DF в xlsx и зарегистрировать как файл.

    Output:
      file_id, name, rows, columns (count), size_bytes,
      size (formatted), sample (first 5 rows), sample_headers (col names)
      — последние 3 нужны UI для превью карточки скачивания.
    """
    df, df_id = await _resolve_df(ctx, df_id)
    # —— ГАРД пустого финального df (§3.8) — В САМОМ НАЧАЛЕ, до get_settings/sanucu —
    from backend.agent.query_spec import is_empty_df  # Lazy: pydantic-safe import
    if is_empty_df(df):
        return ToolResult(ok=False, error=(
            "EMPTY_RESULT: после фильтров/агрегата строк не осталось — нечего "
            "выгружать. Проверь range-порог / категориальное значение / период, "
            "либо честно сообщи пользователю. (df пуст)"))
    cfg = get_settings()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = name or f"agent_{df_id}_{ts}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    out_dir = Path(cfg.files_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / filename
    try:
        # Atomic write + fsync + re-verify, чтобы не отдать битый файл
        # юзеру (NFS на DataLab иногда возвращает из to_excel до полной
        # flush'у; xlsxwriter надёжнее openpyxl на Spark Decimal/Timestamp).
        await asyncio.to_thread(_write_xlsx_safe, df, path, sheet_name)
    except Exception as e:
        return ToolResult(ok=False, error=f"export_excel упал: {e}")
    size = path.stat().st_size
    f = ctx.register_file(name=filename, path=str(path), size_bytes=size,
                          mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Превью для UI: первые 5 строк * все колонки (max 12 для компактности).
    cols = list(df.columns)[:12]
    sample_headers = [str(c) for c in cols]
    sample: list[list] = []
    try:
        for _, row in df.head(5).iterrows():
            sample.append([_safe_cell(row[c]) for c in cols])
    except Exception as e:  # noqa: BLE001
        logger.warning("[export_excel] sample build failed: %s", e)

    return ToolResult(
        ok=True,
        output={"file_id": f.file_id, "name": f.name,
                "rows": len(df), "columns": len(df.columns),
                "size_bytes": size,
                "size": _format_bytes(size),
                "sample": sample,
                "sample_headers": sample_headers},
        summary=f"'{f.file_id}': {f.name} ({len(df)} строк)",
    )


def _normalize_df_for_excel(df):
    """Готовит DataFrame для записи в xlsx — Spark/Hive часто отдаёт типы,
    которые openpyxl/xlsxwriter не пишут корректно:
      * decimal.Decimal -> float (openpyxl пишет Decimal как str "1234.5",
        Excel показывает как текст, не как число)
      * tz-aware datetime -> tz-naive (openpyxl бросает TypeError на tz-aware,
        xlsxwriter тихо пишет битую дату)
      * numpy.int64/float64 -> python int/float
      * Все остальные «странные» object-колонки -> str
    """
    from decimal import Decimal
    import math
    import pandas as pd
    out = df.copy()
    for col in out.columns:
        ser = out[col]
        # tz-aware datetime -> naive
        if hasattr(ser, "dt"):
            try:
                if ser.dt.tz is not None:
                    out[col] = ser.dt.tz_localize(None)
                    continue
            except (TypeError, AttributeError):
                pass
        # Object -> возможно Decimal/Spark Row
        if ser.dtype == object:
            sample = ser.dropna().head(1)
            if len(sample) > 0:
                v = sample.iloc[0]
                if isinstance(v, Decimal):
                    out[col] = ser.apply(
                        lambda x: float(x) if isinstance(x, Decimal)
                                  else (None if x is None else x))
    return out


def _write_xlsx_safe(df, path, sheet_name):
    """Atomic write с fsync и verify:
    1. Нормализуем типы (Decimal -> float, tz-aware -> naive)
    2. Пишем во временный *.tmp файл — engine xlsxwriter (надёжнее
       openpyxl на Spark-типах)
    3. fsync на dir и file — чтобы NFS точно flushнул на disk
    4. Атомарный rename .tmp -> final path
    5. Re-open через openpyxl read-only — verify что xlsx валидный
    6. Если не валидный — удаляем и бросаем исключение
    """
    import os
    import shutil
    from pathlib import Path as _P

    path = _P(path)
    tmp = path.with_suffix(path.suffix + ".tmp")

    # 1. Нормализация
    df_safe = _normalize_df_for_excel(df)

    # 2. Запись в .tmp через xlsxwriter (надёжнее openpyxl на Decimal/NaT)
    try:
        df_safe.to_excel(tmp, sheet_name=sheet_name, index=False,
                         engine="xlsxwriter")
    except ImportError:
        # xlsxwriter не установлен - fallback в openpyxl
        df_safe.to_excel(tmp, sheet_name=sheet_name, index=False,
                         engine="openpyxl")

    # 3. fsync файла + директории — на NFS это критично
    try:
        with open(tmp, "rb") as fh:
            os.fsync(fh.fileno())
        dir_fd = os.open(str(tmp.parent), os.O_RDONLY)
        try:
            os.fsync(dir_fd)
        finally:
            os.close(dir_fd)
    except (OSError, AttributeError):
        # fsync не поддерживается на этой ФС — ладно, пропускаем
        pass

    # 4. Атомарный rename (на POSIX гарантированно атомарный)
    shutil.move(str(tmp), str(path))

    # 5. Verify через openpyxl read-only — открыть и прочесть первую строку
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb[wb.sheetnames[0]]
        # Прочитать первые 2 строки чтобы убедиться что parsing работает
        _ = list(ws.iter_rows(max_row=2, values_only=True))
        wb.close()
    except Exception as e:
        # Файл побит — удаляем чтобы не отдавать юзеру битый
        try:
            path.unlink()
        except OSError:
            pass
        raise RuntimeError(
            f"xlsx-verify failed: {type(e).__name__}: {e}. "
            "Возможно проблема в типах данных Spark (Decimal/Timestamp) "
            "или повреждение при записи на NFS."
        ) from e


def _safe_cell(v) -> str:
    """Сериализуем значение ячейки для UI-preview: NaN->"", datetime->ISO, числа с разрядами."""
    import math
    try:
        if v is None:
            return ""
        # pandas NaN / NaT
        if isinstance(v, float) and math.isnan(v):
            return ""
        # datetime / Timestamp
        if hasattr(v, "isoformat"):
            return v.isoformat(sep=" ", timespec="minutes")[:16]
        # числа: format с разделителем тысяч
        if isinstance(v, (int, float)):
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            return f"{v:,}".replace(",", " ")
        s = str(v)
        return s if len(s) <= 200 else s[:197] + "..."
    except Exception:
        return str(v)[:200]


def _format_bytes(n: int) -> str:
    if n < 1024:
        return f"{n} Б"
    if n < 1024 * 1024:
        return f"{n/1024:.1f} КБ"
    return f"{n/(1024*1024):.1f} МБ"


async def export_csv(ctx, df_id: str, name: Optional[str] = None) -> ToolResult:
    """Сохранить DF в csv."""
    df, df_id = await _resolve_df(ctx, df_id)
    # —— ГАРД пустого финального df (§3.8) — В САМОМ НАЧАЛЕ —
    from backend.agent.query_spec import is_empty_df  # lazy: pydantic-safe import
    if is_empty_df(df):
        return ToolResult(ok=False, error=(
            "EMPTY_RESULT: после фильтров/агрегата строк не осталось — нечего "
            "выгружать. Проверь range-порог / категориальное значение / период, "
            "либо честно сообщи пользователю. (df пуст)"))
    cfg = get_settings()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = name or f"agent_{df_id}_{ts}.csv"
    if not filename.endswith(".csv"):
        filename += ".csv"
    out_dir = Path(cfg.files_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / filename
    await asyncio.to_thread(df.to_csv, path, index=False, encoding="utf-8-sig")
    size = path.stat().st_size
    f = ctx.register_file(name=filename, path=str(path), size_bytes=size,
                          mime_type="text/csv")

    return ToolResult(
        ok=True,
        output={"file_id": f.file_id, "name": f.name,
                "rows": len(df), "size_bytes": size},
        summary=f"'{f.file_id}': {f.name} ({len(df)} строк, csv)",
    )


# —— Composite tool: get_ior_details ----------------------------------


async def get_ior_details(ctx, incdnt_sid: str) -> ToolResult:
    """Полное досье одного ИОР: main + fin_impact + recovery + nonfin + stts.

    Composite tool: 5 query'ев в одном — заменяет план из 5 шагов
    для часто-используемого сценария. Идёт через data_store (Spark/DuckDB).
    """
    store = get_data_store()
    try:
        # to_thread - Spark.sql() блокирующий
        main_df = await asyncio.to_thread(
            store.query, table="d6_base_of_knowledge_ior",
            where={"incdnt_sid": incdnt_sid}, limit=1)
    except Exception as e:  # noqa: BLE001
        return ToolResult(ok=False, error=f"main query упал: {e}")

    if main_df.empty:
        return ToolResult(ok=False, error=f"ИОР {incdnt_sid!r} не найден")
    incdnt_id = int(main_df["incdnt_id"].iloc[0])

    # Related — ВСЕ связаны через incdnt_id (включая fin_impact: реальная
    # витрина fin_impact имеет отдельные incdnt_id (FK) и fin_impact_id (PK);
    # связь с инцидентом идёт по incdnt_id, НЕ по fin_impact_id).
    # ВАЖНО: запросы ПОСЛЕДОВАТЕЛЬНЫЕ, не gather. DuckDB-соединение НЕ
    # потокобезопасно — 4 параллельных to_thread на одном connection дают
    # гонку (один query возвращает None -> 'NoneType has no len()'). Каждый
    # запрос здесь — одна строка инцидента, последовательно это миллисекунды.
    try:
        rec_df = await asyncio.to_thread(
            store.query, table="d6_base_of_knowledge_incident_recovery",
            where={"incdnt_id": incdnt_id}, limit=10_000)
        fin_df = await asyncio.to_thread(
            store.query, table="d6_base_of_knowledge_incident_fin_impact",
            where={"incdnt_id": incdnt_id}, limit=10_000)
        nonfin_df = await asyncio.to_thread(
            store.query, table="d6_base_of_knowledge_incident_nonfin_impact",
            where={"incdnt_id": incdnt_id}, limit=10_000)
        stts_df = await asyncio.to_thread(
            store.query, table="d6_base_of_knowledge_incident_stts_chng",
            where={"incdnt_id": incdnt_id},
            order_by="stts_chng_action_dttm", order_desc=False, limit=1000)
    except Exception as e:  # noqa: BLE001
        return ToolResult(ok=False, error=f"related query упал: {e}")

    df_main = ctx.register_dataframe(main_df, f"main по {incdnt_sid}",
                                     "get_ior_details").df_id
    df_rec = ctx.register_dataframe(rec_df, f"возмещения по {incdnt_sid}",
                                    "get_ior_details").df_id
    df_fin = ctx.register_dataframe(fin_df, f"fin_impact по {incdnt_sid}",
                                    "get_ior_details").df_id
    df_nonfin = ctx.register_dataframe(nonfin_df, f"nonfin по {incdnt_sid}",
                                       "get_ior_details").df_id
    df_stts = ctx.register_dataframe(stts_df, f"история статусов {incdnt_sid}",
                                     "get_ior_details").df_id

    return ToolResult(
        ok=True,
        output={
            "incdnt_sid": incdnt_sid, "incdnt_id": incdnt_id,
            "df_main": df_main, "df_recovery": df_rec,
            "df_fin_impact": df_fin, "df_nonfin": df_nonfin,
            "df_status_history": df_stts,
            "counts": {"main": len(main_df), "recovery": len(rec_df),
                       "fin_impact": len(fin_df), "nonfin": len(nonfin_df),
                       "status_history": len(stts_df)},
        },
        summary=f"досье {incdnt_sid}: {len(rec_df)} возмещ, "
                f"{len(fin_df)} fin_impact, {len(nonfin_df)} nonfin, "
                f"{len(stts_df)} status events",
    )


# —— Register Все ----------------------------------------------------


REGISTRY.register(Tool(
    name="query",
    description=(
        "SELECT из одной таблицы БЗ ИОР с фильтрами. Возвращает df_id. "
        "ВНИМАНИЕ: Запрещено использовать для запросов с фильтрацией по ДАТАМ (периодам) или СУММАМ (потери/возмещения)! "
        "Для любых запросов, где есть даты или деньги, используй ИСКЛЮЧИТЕЛЬНО run_query_spec, иначе выгрузка будет неполной."
    ),
    args_schema={
        "type": "object",
        "properties": {
            "table": {"type": "string", "description": "имя таблицы из schema"},
            "where": {"type": "object", "description": "фильтры"},
            "columns": {"type": "array", "description": "колонки (опционально, default *)"},
            "limit": {"type": "integer", "description": "default _MAX_ROWS_DEFAULT"},
            "order_by": {"type": "string"},
            "order_desc": {"type": "boolean", "default": True},
        },
        "required": ["table"],
    },
    returns="{df_id, rows, columns, sample}",
    run=query, category="data",
))

REGISTRY.register(Tool(
    name="filter_df",
    description="Применить pandas-фильтр (.query()) к существующему df_id. "
                "where = строка типа \\\"incdnt_sum > 1000000 and status != 'Закрыт'\\\".",
    args_schema={
        "type": "object",
        "properties": {
            "df_id": {"type": "string"},
            "where": {"type": "string"},
        },
        "required": ["df_id", "where"],
    },
    returns="{df_id, rows}",
    run=filter_df, category="transform",
))

REGISTRY.register(Tool(
    name="top_n",
    description="Топ-N строк df'а по полю.",
    args_schema={
        "type": "object",
        "properties": {
            "df_id": {"type": "string"},
            "by": {"type": "string"},
            "n": {"type": "integer", "default": 10},
            "ascending": {"type": "boolean", "default": False},
        },
        "required": ["df_id", "by"],
    },
    returns="{df_id, rows}",
    run=top_n, category="transform",
))

REGISTRY.register(Tool(
    name="group_by",
    description="Group by + агрегаты. agg = {column: 'sum'|'count'|'mean'|'max'|'min'}.",
    args_schema={
        "type": "object",
        "properties": {
            "df_id": {"type": "string"},
            "by": {"type": "array"},
            "agg": {"type": "object"},
        },
        "required": ["df_id", "by", "agg"],
    },
    returns="{df_id, rows, columns}",
    run=group_by, category="transform",
))

REGISTRY.register(Tool(
    name="derive_column",
    description=(
        "Вычисляемое поле из существующего. Главное применение — извлечь "
        "месяц/год/квартал из даты ПЕРЕД group_by/window_rank для "
        "помесячных/поквартальных отчётов. op: year|month|quarter|day|"
        "month_num. Пример: из incdnt_entry_dt сделать колонку 'month' "
        "(op=month -> '2025-03'), потом group_by по month."
    ),
    args_schema={
        "type": "object",
        "properties": {
            "df_id": {"type": "string"},
            "source": {"type": "string", "description": "исходная колонка (дата)"},
            "new_column": {"type": "string", "description": "имя новой колонки"},
            "op": {"type": "string", "enum": ["year", "month", "quarter", "day", "month_num"],
                   "default": "month"},
        },
        "required": ["df_id", "source", "new_column"],
    },
    returns="{df_id, rows, columns}",
    run=derive_column, category="transform",
))

REGISTRY.register(Tool(
    name="window_rank",
    description=(
        "Оконное ранжирование внутри групп (SQL ROW_NUMBER/RANK OVER "
        "PARTITION BY ... ORDER BY ...). Для отчётов «ТОП-N внутри каждой "
        "группы»: ТОП-20 ЦПР по количеству В КАЖДОМ месяце, аутсайдеры, "
        "N-й по величине в категории. partition_by=окно (напр. 'month'), "
        "order_by=сортировка (напр. 'incdnt_count'), top_n=оставить rank<=N, "
        "order_desc=true(топ)/false(аутсайдеры). Обычно перед ним group_by."
    ),
    args_schema={
        "type": "object",
        "properties": {
            "df_id": {"type": "string"},
            "partition_by": {"type": "string", "description": "колонка окна (или массив)"},
            "order_by": {"type": "string", "description": "колонка сортировки в окне"},
            "order_desc": {"type": "boolean", "default": True},
            "top_n": {"type": "integer", "description": "оставить только rank<=top_n"},
            "rank_col": {"type": "string", "default": "rank"},
            "method": {"type": "string", "enum": ["row_number", "rank", "dense_rank"],
                       "default": "row_number"},
        },
        "required": ["df_id", "order_by"],
    },
    returns="{df_id, rows, columns}",
    run=window_rank, category="transform",
))

REGISTRY.register(Tool(
    name="join_dfs",
    description="Merge двух df'ов по ключу. how: inner|left|right|outer.",
    args_schema={
        "type": "object",
        "properties": {
            "left_df": {"type": "string"},
            "right_df": {"type": "string"},
            "on": {"type": ["string", "array"]},
            "how": {"type": "string", "default": "inner"},
        },
        "required": ["left_df", "right_df", "on"],
    },
    returns="{df_id, rows}",
    run=join_dfs, category="transform",
))

REGISTRY.register(Tool(
    name="export_excel",
    description="Сохранить df в .xlsx и зарегистрировать как file_id для скачивания юзером. "
                "Используй на ФИНАЛЬНОМ шаге плана.",
    args_schema={
        "type": "object",
        "properties": {
            "df_id": {"type": "string"},
            "name": {"type": "string", "description": "имя файла (опц.)"},
            "sheet_name": {"type": "string", "default": "Отчет_ОпРиски"},
        },
        "required": ["df_id"],
    },
    returns="{file_id, name, rows, size_bytes}",
    run=export_excel, category="export",
))

REGISTRY.register(Tool(
    name="export_csv",
    description="Сохранить df в .csv (10х меньше xlsx, для импорта в другие системы).",
    args_schema={
        "type": "object",
        "properties": {
            "df_id": {"type": "string"},
            "name": {"type": "string"},
        },
        "required": ["df_id"],
    },
    returns="{file_id, name, rows}",
    run=export_csv, category="export",
))

REGISTRY.register(Tool(
    name="get_ior_details",
    description="Получить полное досье ОДНОГО инцидента по SID – main + "
                "возмещения + fin_impact + nonfin_impact + история статусов. "
                "Возвращает 5 df_id (df_main, df_recovery, df_fin_impact, "
                "df_nonfin, df_status_history). Это композитный tool – "
                "заменяет план из 5 query+join шагов.",
    args_schema={
        "type": "object",
        "properties": {
            "incdnt_sid": {"type": "string",
                           "description": "Бизнес-ID, формат EVE-NNNNNNN"},
        },
        "required": ["incdnt_sid"],
    },
    returns="{df_main, df_recovery, df_fin_impact, df_nonfin, "
            "df_status_history, counts}",
    run=get_ior_details, category="data",
))