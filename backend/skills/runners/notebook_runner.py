"""
Запуск ноутбуков навыков.

В mock-режиме (`APP_ENV=local` или `MOCK_NOTEBOOK_EXECUTION=true`) -
генерирует demo Excel + структуры данных (stats, dossier, followups),
совместимые с editorial-restraint UI.

В prod-режиме - **in-process исполнение** notebook'ов:
• Парсит .ipynb, извлекает code-ячейки в порядке следования.
• Исполняет их через exec() в backend-процессе.
• Одна общая SparkSession.builder.getOrCreate() переиспользуется
  между всеми запросами (раньше каждый Papermill-вызов = новый
  Python subprocess = новая JVM = новый SparkSession за 5-10с).
• Ячейка с tag 'parameters' инжектируется: после её исполнения мы
  дописываем `param = value` для каждого params[k]=v, перебивая дефолты.
• Jupyter-магия (!pip, %магия) - пропускается.
• cwd на время exec временно меняется на data/generated_files,
  чтобы `df.to_excel('file.xlsx')` в notebook'е писал прямо туда.
• Глобальный Lock - только одно in-process исполнение одновременно
  (Spark всё равно в одной JVM, параллельность через Spark, а не через
  несколько одновременных notebook'ов).
"""
from __future__ import annotations

import json
import logging
import os
import random
import threading
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

from backend.config import get_settings
from backend.skills.runners.excel_inspector import inspect_excel

logger = logging.getLogger(__name__)

# Глобальный лок: только одно in-process выполнение notebook'а одновременно
# (Spark - единая JVM, нет смысла запускать параллельно).
_exec_lock = threading.Lock()

@dataclass
class ExecutionResult:
    excel_path: Path
    excel_filename: str
    rows: int = 0
    stats: dict = field(default_factory=dict)
    excel_meta: dict = field(default_factory=dict)         # { size, rows, columns, sample }
    dossier: Optional[dict] = None                         # для досье инцидента
    followups: list[dict] = field(default_factory=list)    # [{label, prompt}]
    text: str = ""                                         # короткое описание для UI
    duration_ms: int = 0
    error: Optional[str] = None
    csv_path: Optional[Path] = None                        # CSV-альтернатива (Phase 4)

@dataclass
class Phase:
    """Phase-event от runner.run_phased() - потребляется flow.py для SSE.

    Поля:
      name  - тип события (см. ниже)
      label - человекочитаемая подпись (для UI status step)
      data  - payload (зависит от типа)

    Типы:
      'spark_starting' - стартует Spark / уже идёт инициализация
      'spark_stage'    - прогресс конкретного stage {stage_id, progress_pct,
                         tasks_done, tasks_total, active_stages}
      'count_done'     - известно сколько строк: {total_rows}
      'excel_starting' - начали запись xlsx: {total_rows}
      'excel_progress' - прогресс bytes: {bytes_written}
      'excel_done'     - xlsx готов: {path, size_bytes, csv_path?}
      'result'         - финальный ExecutionResult: {result: ExecutionResult}
      'error'          - упало: {error}
    """

    name: str
    label: str = ""
    data: dict = field(default_factory=dict)

# -------------------------------------------------------------
# Демо-данные для mock (правдоподобные значения)
# -------------------------------------------------------------

TB_OPTIONS = [
    ("СЗБ", "Северо-Западный банк"),
    ("ЮЗБ", "Юго-Западный банк"),
    ("СИБ", "Сибирский банк"),
    ("ВВБ", "Волго-Вятский банк"),
    ("ПВБ", "Поволжский банк"),
    ("УРБ", "Уральский банк"),
    ("СРБ", "Среднерусский банк"),
    ("ДВБ", "Дальневосточный банк"),
    ("МБ", "Московский банк"),
]

TYPE_LVL_1 = [
    "Технические сбои",
    "Ошибка ввода данных",
    "Внешнее мошенничество",
    "Внутреннее мошенничество",
    "Нарушение процесса",
    "Ошибки персонала",
    "Действия третьих лиц",
]

PROCESSES = [
    "Кредитование ФЛ",
    "Кредитные карты",
    "РКО ФЛ",
    "Дебетовые карты",
    "Депозиты ФЛ",
    "Потребительское кредитование",
    "Ипотечное кредитование",
]

STATUSES = ["Утверждён", "Исследование", "Утверждение", "Черновик", "Возмещение", "Закрыт"]

RECOVERY_TYPES = [
    "Восст. техн. сбой",
    "Комп. сотрудника",
    "Страховое",
    "Возврат третьих лиц",
]

NONFIN_KINDS = [
    "Жалобы и обращения клиентов",
    "Освещение в СМИ",
    "Воздействие со стороны регулятора",
    "Утечка, потеря или искажение защищаемой информации",
    "Ущерб репутации",
    "Угроза непрерывности деятельности",
]

class NotebookRunner:
    def __init__(self) -> None:
        self.cfg = get_settings()
        self.output_dir = self.cfg.files_path
        self.output_dir.mkdir(parents=True, exist_ok=True)
        # Активные Papermill subprocess'ы для cancel.
        # --- Phase 4.1: cancel-tracking ----------------------
        # При запросе cancel мы убиваем дочерний ipykernel-subprocess
        # вместе с его JVM - это останавливает Spark job в нём.
        self._active_processes: dict[str, object] = {}  # session_id -> Popen
        self._active_processes_lock = threading.Lock()

        # --- Executor для Papermill-вызовов -----------------
        # Каждый pm.execute_notebook() блокирующий и создаёт subprocess.
        # Запускаем через dedicated single-thread executor чтобы:
        # 1) Не блокировать asyncio event loop
        # 2) Сериализовать запросы (один notebook в один момент времени)
        from concurrent.futures import ThreadPoolExecutor
        self._spark_executor = ThreadPoolExecutor(
            max_workers=1, thread_name_prefix="notebook-worker"
        )

    @property
    def spark_executor(self):
        """Dedicated single-thread executor для всех Papermill-операций.
        flow.py использует loop.run_in_executor(runner.spark_executor, ...).
        """
        return self._spark_executor

    def cancel_session(self, session_id: str) -> bool:
        """
        Cancel-через-Papermill пока упрощённый: нет прямого хука к
        дочернему ipykernel-процессу, поэтому возвращаем False.

        Если реально понадобится - можно перехватить Popen Papermill'а
        через nbclient hooks и убить процесс. Пока не реализовано.
        """
        logger.info(
            f"[NotebookRunner] cancel_session({session_id}) - пока no-op "
            f"(Papermill-subprocess kill требует доработки)",
        )
        return False

    def run(self, *, skill_id: str, notebook_path: Optional[Path],
            params: dict) -> ExecutionResult:
        """Запустить notebook или вернуть mock-результат."""
        started = datetime.now()

        if self.cfg.use_mock_runner or not notebook_path:
            logger.info("[NotebookRunner] mock-mode -> %s", skill_id)
            return self._mock_run(skill_id, params, started)

        try:
            return self._real_run(skill_id, notebook_path, params, started)
        except Exception as e:
            logger.exception("[NotebookRunner] Ошибка %s: %s", skill_id, e)
            return ExecutionResult(
                excel_path=Path(""), excel_filename="",
                error=str(e),
                duration_ms=int((datetime.now() - started).total_seconds() * 1000),
            )

    # --- In-process исполнение notebook'а -------------------------
    def run_phased(
        self,
        *,
        skill_id: str,
        notebook_path: Optional[Path],
        params: dict,
        emit,
        session_id: Optional[str] = None,
    ) -> ExecutionResult:
        """Phased-исполнение с push-callback'ом для real-time UI updates.

        `emit(Phase)` - sync-callback который flow.py пробрасывает в
        asyncio.Queue через call_soon_threadsafe. Так worker-thread
        runner'а может стримить прогресс в async-loop, не блокируя.

        Возвращает ExecutionResult (как раньше) после полного завершения.

        В mock-режиме сразу делегирует в _mock_run и эмитит pseudo-фазы.
        """
        started = datetime.now()
        if self.cfg.use_mock_runner or not notebook_path:
            emit(Phase("spark_starting", label="Mock-runner (без Spark)"))
            result = self._mock_run(skill_id, params, started)
            emit(Phase("excel_done", label="Excel готов",
                       data={
                           "path": str(result.excel_path),
                           "size_bytes": result.excel_path.stat().st_size,
                           "csv_path": str(result.csv_path) if result.csv_path else None
                       }))
            return result

        try:
            return self._real_run_phased(skill_id, notebook_path, params,
                                         started, emit, session_id)
        except Exception as e:
            logger.exception("[NotebookRunner] Ошибка %s: %s", skill_id, e)
            emit(Phase("error", label=f"Ошибка: {e}", data={"error": str(e)}))
            return ExecutionResult(
                excel_path=Path(""), excel_filename="",
                error=str(e),
                duration_ms=int((datetime.now() - started).total_seconds() * 1000),
            )

    # --- Back-compat: старый sync API .run() ---------------------
    def run(self, *, skill_id: str, notebook_path: Optional[Path],
            params: dict) -> ExecutionResult:
        """Совместимость со старым API - не эмитит фаз."""
        return self.run_phased(
            skill_id=skill_id,
            notebook_path=notebook_path,
            params=params,
            emit=lambda _phase: None,
        )

    def _real_run_phased(self, skill_id: str, notebook_path: Path,
                         params: dict, started: datetime,
                         emit, session_id: Optional[str] = None) -> ExecutionResult:
        """Phased-запуск через Papermill в чистом subprocess.

        ⚠️ ИСТОРИЯ. Сначала была версия с in-process exec(): notebook
        выполнялся в backend-процессе через exec() в worker-thread'е,
        с расчётом переиспользовать одну SparkSession между запросами
        (нет JVM startup per request). Идея красивая, но **в реальной
        prod-среде DataLab + Python 3.12 + uvloop + asyncio + pyspark
        3.5.3 - не работает**:
        * JVM gateway мистически умирает через ~10с после создания,
          без stderr-ошибок (не OOM - в DataLab pod'е 128 GB лимит).
        * Подозрение: py4j gateway thread, форкнутый из процесса
          где главный thread под управлением uvloop'а, теряет какой-то
          signal/socket handler и тихо завершается.
        * Несколько раундов фиксов (dedicated thread, убрали
          redirect_stdout, убрали .count(), снизили memory до 1g,
          singleton spark cache) - не помогли.

        Чтобы не утопить весь продукт в этой архитектурной авантюре,
        вернулись к Papermill: каждый chat-запрос запускает свежий
        ipykernel-subprocess, JVM в нём стартует ОДИН раз и работает
        стабильно до конца запроса. Цена: 5-10с JVM startup per request
        (приемлемо). Бонус: чистая изоляция от нашего backend.

        Phased UI остался: ExcelProgressPoller следит за растущим
        xlsx-файлом на диске и эмитит file_progress events.
        """
        try:
            import papermill as pm
        except ImportError:
            raise RuntimeError(
                "papermill не установлен. Запустите pip install -r requirements.txt"
            )

        ts = started.strftime("%Y%m%d_%H%M%S")
        out_nb = self.output_dir / f"{skill_id}_{ts}.ipynb"

        # Snapshot xlsx ДО запуска - найдём новые после
        watch_dirs = [self.output_dir, notebook_path.parent]
        before: set[tuple[str, str]] = set()
        for d in watch_dirs:
            for p in d.glob("*.xlsx"):
                before.add((str(p.parent), p.name))

        logger.info(
            "[NotebookRunner:papermill] %s -> %s, params=%s, cwd=%s",
            notebook_path.name, out_nb.name, params, self.output_dir,
        )

        emit(Phase("spark_starting",
                   label="Запуск notebook (Papermill + Spark)..."))

        # Поллер размера xlsx на диске - даёт UI реальный progress
        excel_progress_poller = _ExcelProgressPoller(
            self.output_dir, emit, interval=0.5,
        )
        excel_progress_poller.start()
        try:
            with _exec_lock:
                pm.execute_notebook(
                    input_path=str(notebook_path),
                    output_path=str(out_nb),
                    parameters=params,
                    kernel_name="python3",
                    cwd=str(self.output_dir),
                    request_save_on_cell_execute=False,
                    log_output=True, # печатаем [NotebookOutput] в uvicorn лог
                    progress_bar=False,
                )
        finally:
            excel_progress_poller.stop()

        # Найти созданные .xlsx после исполнения
        excel_path: Optional[Path] = None
        candidates: list[Path] = []
        for d in watch_dirs:
            for p in d.glob("*.xlsx"):
                key = (str(p.parent), p.name)
                if key not in before:
                    candidates.append(p)
                elif p.stat().st_mtime > started.timestamp():
                    candidates.append(p)

        if candidates:
            # Выбираем самый свежий
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            picked = candidates[0]
            # Если файл не в output_dir - переносим
            if picked.parent != self.output_dir:
                target = self.output_dir / picked.name
                picked.rename(target)
                excel_path = target
            else:
                excel_path = picked

        if excel_path is None or not excel_path.exists():
            raise FileNotFoundError(
                "Excel-файл не появился после исполнения ноутбука "
                f"({notebook_path.name})."
            )

        # --- Phase 4.2: CSV-альтернатива -------------------------
        # Делаем csv-копию для тех, кому нужно открыть в чём-то кроме Excel.
        # На 800к строк csv пишется в ~10х быстрее xlsx, но мы его уже
        # пишем - поэтому здесь это just-in-case fallback us xlsx.
        # Реальная экономия: следующий рантайм можно делать сразу csv (TODO).
        csv_path: Optional[Path] = None
        try:
            import pandas as pd
            df_for_csv = pd.read_excel(excel_path, engine="openpyxl")
            csv_path = excel_path.with_suffix(".csv")
            df_for_csv.to_csv(csv_path, index=False, encoding="utf-8-sig")
            logger.info("[NotebookRunner] CSV: %s (%s)",
                        csv_path.name, _format_bytes(csv_path.stat().st_size))
        except Exception as e:
            logger.warning("[NotebookRunner] CSV-копия не создана: %s", e)
            csv_path = None

        # Парсинг Excel
        info = inspect_excel(excel_path)
        stats = info["stats"]
        excel_meta = info["excel_meta"]

        duration_ms = int((datetime.now() - started).total_seconds() * 1000)
        stats["duration_ms"] = duration_ms

        # Узкие skill-specific обогащения
        followups = _build_followups(skill_id, params)
        narrative = _build_narrative(skill_id, params, stats)
        dossier = None # из реальных данных досье не реконструируем - в БД нет такой структуры

        size_bytes = excel_path.stat().st_size
        emit(Phase("excel_done",
                   label=f"Excel готов • {_format_bytes(size_bytes)}",
                   data={
                       "path": str(excel_path),
                       "size_bytes": size_bytes,
                       "csv_path": str(csv_path) if csv_path else None,
                   }))

        logger.info(
            "[NotebookRunner:real] OK %s rows=%d size=%s duration=%dms",
            skill_id, stats.get("rows", 0),
            excel_meta.get("size", "?"), duration_ms,
        )

        return ExecutionResult(
            excel_path=excel_path,
            excel_filename=excel_path.name,
            rows=stats.get("rows", 0),
            stats=stats,
            excel_meta=excel_meta,
            dossier=dossier,
            followups=followups,
            text=narrative,
            duration_ms=duration_ms,
            csv_path=csv_path,
        )

    # --- Mock-запуск ---------------------------------------------
    def _mock_run(self, skill_id: str, params: dict,
                  started: datetime) -> ExecutionResult:
        import pandas as pd

        ts = started.strftime("%Y%m%d_%H%M%S")
        begin = params.get("incdnt_entry_dt_begin", "2025-01-01")
        end = params.get("incdnt_entry_dt_end", "2025-12-31")
        sid = params.get("incdnt_sid", "EVE-DEMO-001")

        # Filename - соответствует названию реального скрипта
        filename = _filename_for_skill(skill_id, begin, end, sid, ts)
        excel_path = self.output_dir / filename
        # Scale row count with period duration
        try:
            d_begin = datetime.strptime(begin, "%Y-%m-%d")
            d_end = datetime.strptime(end, "%Y-%m-%d")
            days = max(1, (d_end - d_begin).days + 1)
        except Exception:
            days = 90
        rows = max(50, min(100000, int(days * random.uniform(400.0, 600.0))))

        # Excel-данные (sample для preview + полный df для xlsx)
        data_full, sample_preview = _build_excel_rows(skill_id, rows, begin, end, sid)
        df = pd.DataFrame(data_full)
        df.to_excel(excel_path, sheet_name="Отчет_ОпРиски", index=False, engine="openpyxl")
        size_bytes = excel_path.stat().st_size

        stats = _build_stats(skill_id, rows, data_full)
        dossier = _build_dossier(sid) if skill_id == "report_period_specific_ior_v2" else None
        followups = _build_followups(skill_id, params)
        narrative = _build_narrative(skill_id, params, stats)

        duration_ms = int((datetime.now() - started).total_seconds() * 1000)
        stats["duration_ms"] = duration_ms

        logger.info(
            "[NotebookRunner:mock] %s rows=%d size=%s duration=%dms file=%s",
            skill_id, rows, _format_bytes(size_bytes), duration_ms, filename,
        )

        # Генерируем мок-копию CSV для консистентности мок-режима
        csv_path = excel_path.with_suffix(".csv")
        try:
            df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        except Exception:
            csv_path = None

        return ExecutionResult(
            excel_path=excel_path,
            excel_filename=filename,
            rows=rows,
            stats=stats,
            excel_meta={
                "name": filename,
                "rows": rows,
                "size": _format_bytes(size_bytes),
                "columns": len(data_full[0]) if data_full else 0,
                "sample": sample_preview,
            },
            dossier=dossier,
            followups=followups,
            text=narrative,
            duration_ms=duration_ms,
            csv_path=csv_path,
        )

    @staticmethod
    def _count_rows(xlsx_path: Path) -> int:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            return max(0, ws.max_row - 1)
        except Exception:
            return 0

    @staticmethod
    def _count_columns(xlsx_path: Path) -> int:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            return ws.max_column or 0
        except Exception:
            return 0

# --- Helpers --------------------------------------------------------

def _strip_jupyter_magics(src: str) -> str:
    """Убирает строки Jupyter-магии: !shell, %magic, %%cellmagic, ?help.

    Также корректно обрабатывает многострочные команды с
    backslash-continuation:

        !pip install --quiet \\
            --index-url https://... \\
            openpyxl pandas

    - удаляются ВСЕ строки до завершающей (без '\\' на конце), иначе
    оставшиеся продолжения становятся невалидным Python и валят exec().
    """
    keep: list[str] = []
    in_magic_continuation = False
    for line in src.split("\n"):
        stripped = line.rstrip()
        if in_magic_continuation:
            # пропускаем продолжение; если эта строка не заканчивается
            # на '\' - на ней магия завершилась.
            if not stripped.endswith("\\"):
                in_magic_continuation = False
            continue
        s = line.lstrip()
        if s.startswith(('!', '%', '?')):
            if stripped.endswith("\\"):
                in_magic_continuation = True
            continue
        keep.append(line)
    return "\n".join(keep)

def _format_bytes(n: int) -> str:
    if n < 1024:
        return f"{n} Б"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} КБ"
    return f"{n / 1024 / 1024:.1f} МБ"

def _filename_for_skill(skill_id: str, begin: str, end: str,
                        sid: str, ts: str) -> str:
    if skill_id == "report_period_specific_ior_v2":
        return f"Отчёт по ИОР {sid}_{ts}.xlsx"
    if skill_id == "vozmeshenie_ior_v2":
        return f"Возмещения по ИОР {begin} - {end}.xlsx"
    if skill_id == "financial_consequences_ior_v2":
        return f"Финансовые последствия ИОР {begin} - {end}.xlsx"
    if skill_id == "ior_nonfinancial_consequences_v2":
        return f"Нефинансовые последствия ИОР {begin} - {end}.xlsx"
    if skill_id == "deleted_ior_v2":
        return f"Удалённые ИОР за период {begin} - {end}.xlsx"
    if skill_id == "ior_hypothesis_v2":
        return f"Выгрузка ИОР для анализа гипотез {begin} - {end}.xlsx"
    return f"ИОР за период по ПАО Сбербанк {begin} - {end}.xlsx"

def _build_excel_rows(skill_id: str, rows: int, begin: str,
                      end: str, sid: str) -> tuple[list[dict], list[list]]:
    """Возвращает (data_full_for_df, preview_sample_5x6)."""
    data: list[dict] = []
    # Будем формировать колонки в зависимости от типа отчёта
    base_date = begin
    is_dossier = skill_id == "report_period_specific_ior_v2"
    use_full_67 = skill_id in ("ior_hypothesis_v2", "ior_period_pao_sberbank_v2", "deleted_ior_v2")

    for i in range(rows):
        tb_short, tb_full = random.choice(TB_OPTIONS)
        eve = sid if (is_dossier and i < 3) else f"EVE-{5_000_000 + random.randint(0, 999_999)}"
        d = _random_date(begin, end)
        type_l1 = random.choice(TYPE_LVL_1)
        process = random.choice(PROCESSES)
        amount = round(random.uniform(50_000, 5_000_000), 2) if random.random() > 0.3 else None
        status = random.choice(STATUSES)
        autoreg = random.choice(["Y", "N"])

        if use_full_67:
            row = {
                "Идентификационный ключ инцидента операционного риска": random.randint(10**18, 10**19 - 1),
                "Идентификатор события": eve,
                "Статус события": status,
                "Признак авторегистрации инцидента": autoreg,
                "Кем выявлено событие": "Система мониторинга",
                "Название источника": "РКО ФЛ",
                "Тип источника инцидента (уровень 1)": "Внутренняя система",
                "Тип источника инцидента (уровень 2)": "Сервер авторизации",
                "Тип события – уровень 1": type_l1,
                "Тип события – уровень 2": "Технический сбой",
                "Дата обнаружения (Событие)": d,
                "Дата начала инцидента операционного риска": d,
                "Дата ввода (Событие)": d,
                "Дата первого подтверждения": d,
                "Дата последнего подтверждения": d,
                "Идентификатор профиля риска": "RP-08",
                "Наименование профиля риска": "Информационная безопасность",
                "Тип клиента": "Физическое лицо",
                "Количество ошибок": random.randint(1, 5),
                "Номер заявки": "REQ-" + str(random.randint(100000, 999999)),
                "Номер договора": "AGR-" + str(random.randint(100000, 999999)),
                "Идентификатор договора": "AGR-SID-" + str(random.randint(100000, 999999)),
                "Предварительное описание": "Описание инцидента операционного риска",
                "Подробное описание": "Подробное описание инцидента операционного риска",
                "Идентификатор оргструктуры": "ORG-" + str(random.randint(1000, 9999)),
                "Орг. структура – уровень 2 (Терр. структура / Департамент ДЗО)": "ПАО Сбербанк",
                "Орг. структура – уровень 3 (Блок / ТБ / ПЦП)": tb_full,
                "Орг. структура – уровень 4 (Дивизион / Департамент)": "Отделение",
                "Орг. структура – уровень 5": None,
                "Орг. структура – уровень 6": None,
                "Орг. структура – уровень 7": None,
                "Орг. структура – уровень 8": None,
                "Орг. структура – уровень 9": None,
                "Орг. структура – уровень 10": None,
                "Идентификатор функционального блока": "FB-" + str(random.randint(100, 999)),
                "Функциональный блок – уровень 2": "Розничный бизнес",
                "Функциональный блок – уровень 3": "Депозиты и расчеты",
                "Функциональный блок – уровень 4": "РКО ФЛ",
                "Процесс – уровень 1": "РКО",
                "Процесс – уровень 2": "РКО ФЛ",
                "Процесс – уровень 3": "Проведение платежа",
                "Процесс – уровень 4 (Наименование процесса)": process,
                "Клиентский путь – уровень 4": "Платежи и переводы",
                "Идентификационный ключ направления деятельности": "BA-" + str(random.randint(100, 999)),
                "Направление деятельности банка": "Розничные банковские услуги",
                "Поднаправление деятельности банка": "РКО ФЛ",
                "Связь с ИБ-риском": random.choice(["Y", "N"]),
                "Связь с риском информационных систем": random.choice(["Y", "N"]),
                "Связь с поведенческим риском": random.choice(["Y", "N"]),
                "Связь с модельным риском": random.choice(["Y", "N"]),
                "Общая сумма всех последствий (руб.)": amount,
                "Прямая потеря – итого (руб.)": amount,
                "Прямая потеря – с кредитным риском (руб.)": None,
                "Прямая потеря – без кредитного риска (руб.)": amount,
                "Косвенная потеря – итого (руб.)": None,
                "Косвенная потеря – с кредитным риском (руб.)": None,
                "Косвенная потеря – без кредитного риска (руб.)": None,
                "Нереализовавшаяся потеря – итого (руб.)": None,
                "Нереализовавшаяся потеря – с кредитным риском (руб.)": None,
                "Нереализовавшаяся потеря – без кредитного риска (руб.)": None,
                "Потеря третьих лиц – итого (руб.)": None,
                "Потеря третьих лиц – с кредитным риском (руб.)": None,
                "Потеря третьих лиц – без кредитного риска (руб.)": None,
                "Прибыль – итого (руб.)": None,
                "Прибыль – с кредитным риском (руб.)": None,
                "Прибыль – без кредитного риска (руб.)": None,
                "Возмещение – итого по инциденту (руб.)": round(amount * random.uniform(0.1, 0.4), 2) if amount else None,
            }
        else:
            row = {
                "Идентификатор события": eve,
                "Дата ввода": d,
                "ТБ": tb_full,
                "Тип события - уровень 1": type_l1,
                "Сумма последствий, ₽": amount,
                "Статус": status,
                "Процесс": process,
                "Авторегистрация": autoreg,
            }

        if skill_id == "vozmeshenie_ior_v2":
            row["Тип возмещения"] = random.choice(RECOVERY_TYPES)
            row["Сумма возмещения, ₽"] = round(random.uniform(10_000, 1_500_000), 2)
        elif skill_id == "financial_consequences_ior_v2":
            row["Тип последствия"] = random.choice([
                "Прямая потеря", "Косвенная потеря", "Нереализовавшаяся",
                "Потеря третьих лиц", "Прибыль",
            ])
        elif skill_id == "ior_nonfinancial_consequences_v2":
            row["Вид качественной потери"] = random.choice(NONFIN_KINDS)
            row["Класс влияния"] = random.choice([
                "Низкий", "Средний", "Высокий", "Очень высокий"
            ])
        elif skill_id == "deleted_ior_v2":
            if use_full_67:
                row["Статус события"] = "Удалён"
            else:
                row["Статус"] = "Удалён"
            row["Причина удаления"] = random.choice([
                "дубликат", "не является ОР", "ошибка регистрации",
                "уточнение от риск-координатора",
            ])

        data.append(row)

    # Preview: первые 5 рядов в формате 6-колонок для UI (SID, Дата, ТБ, Тип, Сумма, Статус)
    preview: list[list] = []
    for r in data[:5]:
        amt = r.get("Общая сумма всех последствий (руб.)") or r.get("Сумма последствий, ₽")
        amt_str = f"{amt:,.0f} ₽".replace(",", " ") if amt is not None else "—"
        
        eve = r.get("Идентификатор события")
        dt = r.get("Дата ввода (Событие)") or r.get("Дата ввода")
        tb = r.get("Орг. структура – уровень 3 (Блок / ТБ / ПЦП)") or r.get("ТБ") or ""
        type_val = r.get("Тип события – уровень 1") or r.get("Тип события - уровень 1") or ""
        status_val = r.get("Статус события") or r.get("Статус") or ""
        
        preview.append([
            eve,
            dt,
            tb.replace("банк", "Б") if "Северо" in tb else tb,
            type_val,
            amt_str,
            status_val,
        ])
    return data, preview

def _build_stats(skill_id: str, rows: int, data: list[dict]) -> dict:
    """Структура совместимая с UI: top_tb/type/process - объекты, breakdown - массивы."""
    from collections import Counter

    tbs = []
    types = []
    procs = []
    sum_total = 0.0
    n_autoreg = 0

    for r in data:
        tb = r.get("Орг. структура – уровень 3 (Блок / ТБ / ПЦП)") or r.get("ТБ") or ""
        tbs.append(tb)

        type_val = r.get("Тип события – уровень 1") or r.get("Тип события - уровень 1") or ""
        types.append(type_val)

        proc = r.get("Процесс – уровень 4 (Наименование процесса)") or r.get("Процесс") or ""
        procs.append(proc)

        amount = r.get("Общая сумма всех последствий (руб.)") or r.get("Сумма последствий, ₽") or 0.0
        sum_total += float(amount)

        ar = r.get("Признак авторегистрации инцидента") or r.get("Авторегистрация")
        if ar == "Y":
            n_autoreg += 1

    tb_counter = Counter(tbs)
    type_counter = Counter(types)
    proc_counter = Counter(procs)

    top_tb_label, top_tb_value = tb_counter.most_common(1)[0] if tbs else ("", 0)
    top_type_label, top_type_value = type_counter.most_common(1)[0] if types else ("", 0)
    top_proc_label, top_proc_value = proc_counter.most_common(1)[0] if procs else ("", 0)

    sum_recovery = round(sum_total * random.uniform(0.18, 0.42), 2)

    breakdown_type = [
        {"label": label, "value": v}
        for label, v in type_counter.most_common(5)
    ]

    # Распределение по месяцам - синтезируем
    breakdown_month = [
        {"label": f"{m:02d}", "value": random.randint(rows // 18, rows // 8)}
        for m in range(1, 13)
    ]

    return {
        "rows": rows,
        "n_unique_incdnt_sid": rows,
        "sum_total_loss": round(sum_total, 2),
        "recovery": sum_recovery,
        "n_autoreg": n_autoreg,
        "top_tb": {"label": top_tb_label, "value": top_tb_value},
        "top_type": {
            "label": top_type_label,
            "value": top_type_value,
            "pct": round(top_type_value / rows * 100) if rows else 0,
        },
        "top_process": {
            "label": top_proc_label,
            "value": top_proc_value,
            "pct": round(top_proc_value / rows * 100) if rows else 0,
        },
        "breakdown_type": breakdown_type,
        "breakdown_month": breakdown_month,
    }

def _build_dossier(sid: str) -> dict:
    direct = round(random.uniform(200_000, 3_000_000), 2)
    indirect = round(random.uniform(50_000, 800_000), 2)
    recovery = round(direct * random.uniform(0.6, 0.95), 2)
    tb_short, tb_full = random.choice(TB_OPTIONS)
    return {
        "sid": sid,
        "title": "Двойное списание по кредитным договорам ФЛ из-за сбоя авторизации",
        "status": "Закрыт",
        "entry_dt": "14.03.2025",
        "detection_dt": "12.03.2025",
        "start_dt": "12.03.2025",
        "autoreg": True,
        "risk_profile": "Профиль 8 - Информационная безопасность",
        "type": "Операционные ошибки -> Технические сбои",
        "source": "Система мониторинга -> Автоматическая регистрация",
        "tb": tb_full,
        "func_block": "Розничный бизнес -> Кредитование ФЛ",
        "process": "Кредитование ФЛ -> Выдача -> Авторизация платежа",
        "client_type": "Физическое лицо",
        "summary": (
            "В период 13:42-14:08 произошёл сбой в системе авторизации платежей. "
            "Из-за повторной отправки запросов произошло двойное списание "
            "по 14 кредитным договорам ФЛ. Инцидент закрыт после возмещения клиентам."
        ),
        "amounts": {
            "direct": direct,
            "indirect": indirect,
            "unrealized": 0,
            "third_party": 0,
            "gain": 0,
            "recovery": recovery,
        },
        "fin_impacts": [
            {"type": "Прямая потеря", "kind": "Кредитная", "amount": direct},
            {"type": "Косвенная", "kind": "Некредитная", "amount": indirect},
        ],
        "recoveries": [
            {"type": "Восстановление по техн. сбою", "date": "28.03.2025", "amount": recovery},
        ],
        "flags": {"ib": True, "is": True, "behavior": False, "model": False},
        "timeline": [
            {"label": "Обнаружен", "date": "12.03.2025", "state": "done"},
            {"label": "Зарегистрирован", "date": "14.03.2025", "state": "done"},
            {"label": "На проверке", "date": "18.03.2025", "state": "done"},
            {"label": "Возмещение", "date": "28.03.2025", "state": "done"},
            {"label": "Закрыт", "date": "02.04.2025", "state": "current"},
        ],
        "links": {
            "agr_num": str(random.randint(1_000_000_000, 9_999_999_999)),
            "appl_num": f"ETSM-2025-{random.randint(10000, 99999)}",
        },
    }

def _build_followups(skill_id: str, params: dict) -> list[dict]:
    period = (
        f"{params.get('incdnt_entry_dt_begin', '2025')}-"
        f"{params.get('incdnt_entry_dt_end', '2025')}"
    )
    sid = params.get("incdnt_sid", "EVE-...")
    if skill_id == "report_period_specific_ior_v2":
        return [
            {"label": "Нефин. последствия по " + sid,
             "prompt": f"Покажи нефинансовые последствия по {sid}"},
            {"label": "Виновный сотрудник",
             "prompt": f"Кто работал с {sid}?"},
            {"label": "Похожие ИОР",
             "prompt": f"Найди похожие ИОР по техн. сбоям"},
        ]
    if skill_id == "vozmeshenie_ior_v2":
        return [
            {"label": "Расширить до полугодия", "prompt": "Возмещения за H1 2025"},
            {"label": "По ФИО виновных", "prompt": "Покажи виновных по этим возмещениям"},
        ]
    if skill_id == "deleted_ior_v2":
        return [
            {"label": "Контроль УВА", "prompt": "Удаления после утверждения за этот период"},
            {"label": "По пользователям", "prompt": "Кто чаще всех удалял ИОР?"},
        ]
    return [
        {"label": "Топ-10 по сумме",
         "prompt": "Покажи топ-10 ИОР за этот период по сумме"},
        {"label": "Только тех. сбои",
         "prompt": "Отфильтруй – только технические сбои"},
        {"label": "Сравни с прошлым годом",
         "prompt": "Сравни статистику с прошлым годом"},
    ]

def _build_narrative(skill_id: str, params: dict, stats: dict) -> str:
    begin = params.get("incdnt_entry_dt_begin", "-")
    end = params.get("incdnt_entry_dt_end", "-")
    rows = stats.get("rows", 0)
    # ВАЖНО: stats.get(key, {}) даёт {} только если key отсутствует;
    # если key=None - вернёт None. Поэтому `or {}` – финальный страж
    # от 'NoneType' has no attribute 'get'.
    tt = stats.get("top_type") or {}
    tp = stats.get("top_process") or {}
    sid = params.get("incdnt_sid")

    sum_loss = stats.get("sum_total_loss") or 0
    recov = stats.get("recovery") or 0
    recov_pct = round(recov / sum_loss * 100) if sum_loss else 0

    if skill_id == "report_period_specific_ior_v2" and sid:
        return (
            f"Полное досье по инциденту **{sid}** ниже. По ЦЛР отнесён к "
            f"**Информационной безопасности**, виновный – внешняя система. "
            f"Возмещение получено полностью."
        )

    if skill_id == "vozmeshenie_ior_v2":
        return (
            f"За **{begin} – {end}** найдено **{rows}** операций возмещения на сумму "
            f"**{sum_loss:,.0f} ₽**.".replace(",", " ")
        )

    return (
        f"Готова выгрузка по запросу за период **{begin} – {end}**. \n"
        f"В период попало **{rows}** инцидентов.\n"
        f"Общая сумма потерь — **{sum_loss:,.0f} ₽**, возмещения — **{recov:,.0f} ₽** "
        f"({recov_pct}%).\n\n"
        f"Доминирующий тип — **{tt.get('label','-')}** ({tt.get('pct',0)}%), "
        f"главный процесс — **{tp.get('label','-')}**.\n\n"
        f"📁 Полная выгрузка в Excel."
    ).replace(",", " ")

def _random_date(begin: str, end: str) -> str:
    """Случайная дата в формате DD.MM.YYYY между begin и end."""
    from datetime import date
    try:
        d1 = date.fromisoformat(begin)
        d2 = date.fromisoformat(end)
    except ValueError:
        return begin
    if d2 < d1:
        d1, d2 = d2, d1
    delta = (d2 - d1).days
    rd = d1 if delta == 0 else d1.replace() + (d2 - d1) * 0  # placeholder
    from datetime import timedelta
    rd = d1 + timedelta(days=random.randint(0, max(0, delta)))
    return rd.strftime("%d.%m.%Y")

# --- Progress pollers (background threads) --------------------------

class _ExcelProgressPoller:
    """Опрашивает размер растущего .xlsx файла на диске и эмитит
    Phase('excel_progress', ...) каждые `interval` сек.

    Берёт самый свежий .xlsx с mtime > now-300c в output_dir.
    Так UI видит «записано 12 МБ» в реальном времени, а не ждёт
    минутами пока xlsxwriter закроет файл.
    """

    def __init__(self, output_dir: Path, emit, interval: float = 0.5) -> None:
        self._output_dir = output_dir
        self._emit = emit
        self._interval = interval
        self._stop = threading.Event()
        self._thread: Optional[threading.Thread] = None
        self._known_before: set[str] = {
            p.name for p in output_dir.glob("*.xlsx")
        }

    def start(self) -> None:
        self._thread = threading.Thread(
            target=self._run, name="excel-progress-poller", daemon=True
        )
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()
        if self._thread is not None:
            self._thread.join(timeout=2.0)

    def _run(self) -> None:
        last_bytes = -1
        while not self._stop.wait(self._interval):
            try:
                candidates = [
                    p for p in self._output_dir.glob("*.xlsx")
                    if p.name not in self._known_before
                ]
                if not candidates:
                    continue
                target = max(candidates, key=lambda p: p.stat().st_mtime)
                size = target.stat().st_size
                if size == last_bytes:
                    continue
                last_bytes = size
                self._emit(Phase(
                    "excel_progress",
                    label=f"Запись Excel: {_format_bytes(size)}",
                    data={"bytes_written": size, "name": target.name},
                ))
            except Exception as e:  # noqa: BLE001
                logger.debug("[ExcelPoller] %s", e)
                continue

# --- Singleton ----------------------------------------------

_runner: Optional[NotebookRunner] = None


def get_runner() -> NotebookRunner:
    global _runner
    if _runner is None:
        _runner = NotebookRunner()
    return _runner