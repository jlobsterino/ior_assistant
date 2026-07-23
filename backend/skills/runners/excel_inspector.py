"""
Универсальный парсер Excel-отчётов после реального запуска ноутбуков.

Извлекает stats (top_tb / top_type / top_process / суммы / breakdown) и
preview-sample 5*6 для UI ExcelAttachment – независимо от точных русских
названий колонок (используется паттерн-маппинг).
"""

from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


# --- Паттерны имён колонок (русский, как в RENAME у Зенковского) ---


def _find_col(columns: list[str], *substrings: str) -> Optional[str]:
    """Находит первую колонку, чьё имя содержит ВСЕ подстроки (case-insensitive)."""
    low = {c: c.lower() for c in columns}
    for col, lc in low.items():
        if all(s.lower() in lc for s in substrings):
            return col
    return None


def _first_existing(columns: list[str], *names: str) -> Optional[str]:
    """Возвращает первую существующую колонку из списка alternative."""
    for n in names:
        if n in columns:
            return n
    return None


def inspect_excel(xlsx_path: Path, *, sheet_name: str = "Отчет_ОпРиски", max_sample: int = 5) -> dict:
    """Открывает Excel и возвращает {stats, excel_meta}."""
    # Стратегия:
    # - Пытаемся читать через pandas (быстрее и удобнее агрегировать).
    # - Если pandas не доступен – fallback на openpyxl.
    try:
        import pandas as pd
        return _inspect_pandas(xlsx_path, sheet_name=sheet_name, max_sample=max_sample)
    except ImportError:
        return _inspect_openpyxl(xlsx_path, max_sample=max_sample)
    except Exception as e:
        logger.exception("[ExcelInspector] %s – pandas: %s", xlsx_path.name, e)
        return _inspect_openpyxl(xlsx_path, max_sample=max_sample)


def _inspect_pandas(xlsx_path: Path, *, sheet_name: str = "Отчет_ОпРиски", max_sample: int = 5) -> dict:
    import pandas as pd

    # Сначала считываем заголовки Excel для динамического определения ID-колонок и сохранения их строкового типа
    dtype_dict = {}
    try:
        try:
            df_headers = pd.read_excel(xlsx_path, sheet_name=sheet_name, nrows=0, engine="openpyxl")
        except Exception:
            df_headers = pd.read_excel(xlsx_path, sheet_name=0, nrows=0, engine="openpyxl")
        
        for col in df_headers.columns:
            col_lower = str(col).lower()
            if any(x in col_lower for x in ("id", "sid", "key", "номер", "идентификатор")):
                if any(x in col_lower for x in ("cnt", "sum", "amt", "val", "кол", "кол-во", "сумма")):
                    continue
                dtype_dict[col] = str
    except Exception:
        pass

    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=dtype_dict, engine="openpyxl")
    except Exception:
        df = pd.read_excel(xlsx_path, sheet_name=0, dtype=dtype_dict, engine="openpyxl")

    cols = list(df.columns)
    rows = len(df)

    # --- Маппинг колонок ---
    col_sid = _find_col(cols, "Идентификатор события") \
        or _find_col(cols, "Идентификационный ключ", "инцидента") \
        or _find_col(cols, "incdnt_sid")
    col_date = _find_col(cols, "Дата ввода") or _find_col(cols, "incdnt_entry_dt")
    # Тб: ищем lvl_2 (Терр. структура) ПЕРВЫМ – у lvl_3 в новой схеме
    # часто слишком мелкая категория ("Блок 'Сеть продаж'") с уникальными
    # значениями, что даёт мусорный top.
    col_tb = _find_col(cols, "уровень 2") \
        or _find_col(cols, "Терр", "структура") \
        or _find_col(cols, "уровень 3") \
        or _find_col(cols, "org_struct_lvl_2") \
        or _find_col(cols, "org_struct_lvl_3")
    col_type = _find_col(cols, "Тип события", "уровень 1") \
        or _find_col(cols, "тип", "уровень 1") \
        or _find_col(cols, "incdnt_type_lvl_1")
    col_status = _find_col(cols, "Статус события") or _find_col(cols, "Статус")
    col_process = _find_col(cols, "Процесс", "уровень 4") \
        or _find_col(cols, "process_lvl_4")
    col_amount = _find_col(cols, "Общая сумма", "последств") \
        or _find_col(cols, "Сумма последствий") \
        or _find_col(cols, "сумма последствия") \
        or _find_col(cols, "сумма последств") \
        or _find_col(cols, "incdnt_sum") \
        or _find_col(cols, "Сумма возмещения") \
        or _find_col(cols, "recovery")
    col_recovery = _find_col(cols, "Возмещение") or _find_col(cols, "recovery_rub_amt_aggr")
    col_autoreg = _find_col(cols, "авторегистр")

    # --- Stats – top values + breakdowns ---
    def _top(col: Optional[str]) -> Optional[dict]:
        if not col or rows == 0:
            return None
        s = df[col].dropna()
        if s.empty:
            return None
        top = Counter(s).most_common(1)[0]
        val = top[1]
        # Если top value == 1 при rows > 50 – значит колонка содержит
        # уникальные значения (например, описание ИОР), это не настоящая
        # категория. Возвращаем None чтобы UI показал «-» вместо мусора.
        if val == 1 and rows > 50:
            return None
        return {
            "label": str(top[0]),
            "value": int(val),
            "pct": round(val / rows * 100),
        }

    def _breakdown(col: Optional[str], n: int = 5) -> Optional[list[dict]]:
        if not col or rows == 0:
            return None
        s = df[col].dropna()
        if s.empty:
            return None
        return [
            {"label": str(k), "value": int(v)}
            for k, v in Counter(s).most_common(n)
        ]

    def _sum(col: Optional[str]) -> float:
        if not col or rows == 0:
            return 0.0
        try:
            return float(df[col].fillna(0).sum())
        except Exception:
            return 0.0

    sum_total = _sum(col_amount)
    sum_recovery = _sum(col_recovery)

    # Breakdown по месяцам (если есть дата) – парсим
    breakdown_month: Optional[list[dict]] = None
    if col_date and rows > 0:
        try:
            dates = pd.to_datetime(df[col_date], errors="coerce", dayfirst=True).dropna()
            if not dates.empty:
                months = dates.dt.month
                cnt = Counter(months)
                breakdown_month = [
                    {"label": f"{m:02d}", "value": int(cnt.get(m, 0))}
                    for m in range(1, 13)
                ]
        except Exception as e:
            logger.debug("[ExcelInspector] breakdown_month skip: %s", e)

    stats: dict[str, Any] = {
        "rows": rows,
        "n_unique_incdnt_sid": int(df[col_sid].nunique()) if col_sid else rows,
        "sum_total_loss": round(sum_total, 2),
        "recovery": round(sum_recovery, 2),
        "top_tb": _top(col_tb),
        "top_type": _top(col_type),
        "top_process": _top(col_process),
        "breakdown_type": _breakdown(col_type),
        "breakdown_month": breakdown_month,
    }
    if col_autoreg:
        s = df[col_autoreg].dropna().astype(str).str.upper()
        stats["n_autoreg"] = int((s == "Y").sum())

    # --- Sample 5x6 для UI ---
    sample = _build_sample(
        df, max_sample=max_sample,
        col_sid=col_sid, col_date=col_date, col_tb=col_tb,
        col_type=col_type, col_amount=col_amount, col_status=col_status,
    )

    # --- Заголовки колонок для превью ---
    matched_count = sum(1 for c in [col_sid, col_date, col_tb, col_type, col_amount, col_status] if c is not None)

    if matched_count < 3:
        # fallback в _build_sample – первые 6 колонок as-is
        sample_headers = [str(c) for c in list(cols)[:6]]
    else:
        sample_headers = ["SID", "Дата", "ТБ", "Тип", "Сумма", "Статус"]

    excel_meta = {
        "name": xlsx_path.name,
        "rows": rows,
        "columns": len(cols),
        "size": _format_bytes(xlsx_path.stat().st_size),
        "sample": sample,
        "sample_headers": sample_headers,
    }
    return {"stats": stats, "excel_meta": excel_meta}


def _build_sample(df, *, max_sample: int, col_sid, col_date, col_tb,
                  col_type, col_amount, col_status) -> list[list]:
    """5x6 превью для ExcelAttachment.

    Если pattern-matching не нашёл хотя бы 3 из 6 "канонических" колонок
    (SID/дата/ТБ/тип/сумма/статус) – fallback: показываем первые 6 колонок
    DataFrame'а as-is. Юзер хотя бы увидит реальные данные, а не пачку «-».
    """
    import pandas as pd

    matched = [c for c in [col_sid, col_date, col_tb, col_type, col_amount, col_status] if c is not None]
    if len(matched) < 3:
        # Fallback: первые 6 колонок DataFrame'а как есть
        sample: list[list] = []
        head_cols = list(df.columns)[:6]
        for _, r in df.head(max_sample).iterrows():
            row = []
            for c in head_cols:
                v = r[c]
                if pd.isna(v):
                    row.append("-")
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    # Числа – форматируем
                    if isinstance(v, float):
                        row.append(f"{v:,.2f}".replace(",", " "))
                    else:
                        row.append(f"{v:,}".replace(",", " "))
                else:
                    row.append(str(v)[:50])  # не больше 50 символов
            sample.append(row)
        return sample

    # Канонический путь – у нас есть нужные колонки
    sample: list[list] = []
    head = df.head(max_sample)
    for _, r in head.iterrows():
        sid = str(r[col_sid]) if col_sid and not pd.isna(r[col_sid]) else "-"
        date_v = r[col_date] if col_date and not pd.isna(r[col_date]) else None
        if isinstance(date_v, datetime):
            date_str = date_v.strftime("%d.%m.%Y")
        elif date_v is not None:
            try:
                date_str = pd.to_datetime(date_v, dayfirst=True).strftime("%d.%m.%Y")
            except Exception:
                date_str = str(date_v)
        else:
            date_str = "-"
        tb_str = str(r[col_tb]) if col_tb and not pd.isna(r[col_tb]) else "-"
        type_str = str(r[col_type]) if col_type and not pd.isna(r[col_type]) else "-"
        amt = r[col_amount] if col_amount and not pd.isna(r[col_amount]) else None
        amt_str = (
            f"{float(amt):,.0f} ₽".replace(",", " ")
            if amt is not None else "-"
        )
        st = str(r[col_status]) if col_status and not pd.isna(r[col_status]) else "-"
        sample.append([sid, date_str, tb_str, type_str, amt_str, st])
    return sample


def _inspect_openpyxl(xlsx_path: Path, *, max_sample: int = 5) -> dict:
    """Fallback без pandas – только rows/columns/size, без stats."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        rows = max(0, ws.max_row - 1)
        cols = ws.max_column or 0
        sample: list[list] = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=1 + max_sample, values_only=True)):
            sample.append([str(c) if c is not None else "-" for c in row[:6]])
            if i >= max_sample - 1:
                break
        return {
            "stats": {"rows": rows, "duration_ms": 0},
            "excel_meta": {
                "name": xlsx_path.name,
                "rows": rows,
                "columns": cols,
                "size": _format_bytes(xlsx_path.stat().st_size),
                "sample": sample,
            },
        }
    except Exception as e:
        logger.exception("[ExcelInspector] openpyxl fallback failed: %s", e)
        return {
            "stats": {"rows": 0},
            "excel_meta": {
                "name": xlsx_path.name,
                "rows": 0,
                "columns": 0,
                "size": "-",
                "sample": [],
            }
        }


def _format_bytes(n: int) -> str:
    if n < 1024:
        return f"{n} Б"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} КБ"
    return f"{n / 1024 / 1024:.1f} МБ"