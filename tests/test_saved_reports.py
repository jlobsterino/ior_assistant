"""Регрессия saved_reports (П5) - хранилище именованных отчетов + лист методологии."""
import os
import tempfile

from backend.agent import saved_reports as sr
from backend.agent.result import write_methodology_sheet

_SPEC = {"source": {"table": "d6_base_of_knowledge_ior"},
         "filters": [{"kind": "period", "intent": {"text": "Q1 2026"}}],}


def _with_store(fn):
    path = tempfile.mktemp(suffix=".json")
    os.environ["REPORTS_STORE"] = path
    try:
        fn()
    finally:
        os.environ.pop("REPORTS_STORE", None)
        if os.path.exists(path):
            os.remove(path)


def test_save_list_get_delete():
    def body():
        r = sr.save_report("ИОР ВВБ Q1", _SPEC, query="ИОР по ВВБ за Q1 2026", now=1.0)
        assert r["id"] and r["name"] == "ИОР ВВБ Q1"
        lst = sr.list_reports()
        assert len(lst) == 1 and lst[0]["query"] == "ИОР по ВВБ за Q1 2026"
        got = sr.get_report(lst[0]["id"])
        assert got["spec"] == _SPEC
        assert sr.delete_report(lst[0]["id"]) is True
        assert sr.list_reports() == []
    _with_store(body)


def test_save_overwrites_by_name():
    def body():
        sr.save_report("Отчёт", {"v": 1}, now=1.0)
        sr.save_report("Отчёт", {"v": 2}, now=2.0)
        lst = sr.list_reports()
        assert len(lst) == 1   # без дублей по имени
        assert sr.get_report(lst[0]["id"])["spec"] == {"v": 2}
    _with_store(body)


def test_methodology_sheet_written():
    import pandas as pd
    from openpyxl import load_workbook
    p = tempfile.mktemp(suffix=".xlsx")
    pd.DataFrame({"a": [1]}).to_excel(p, index=False)
    ok = write_methodology_sheet(p, _SPEC, funnel=[{"stage": "Инциденты", "rows": 100}])
    assert ok
    wb = load_workbook(p)
    assert "Методология" in wb.sheetnames
    assert wb["Методология"]["A1"].value == "Методология выгрузки"
    os.remove(p)


if __name__ == "__main__":
    import traceback
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    passed = 0
    for fn in fns:
        try:
            fn(); print(f" ok  {fn.__name__}"); passed += 1
        except Exception:
            print(f"FAIL {fn.__name__}")
            traceback.print_exc()
    print(f"\n{passed}/{len(fns)} passed")
    raise SystemExit(0 if passed == len(fns) else 1)